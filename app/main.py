from __future__ import annotations

import os
import io
import json
import re
import uuid
import tempfile
from copy import deepcopy
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
from typing import Any, Dict, List, Optional, Tuple

import polars as pl
from fastapi import FastAPI, Form, Request, UploadFile, File, HTTPException
from fastapi.responses import HTMLResponse, StreamingResponse, RedirectResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import jinja2
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

try:
    import pandas as pd  # optional for xlsx
except Exception:
    pd = None  # type: ignore

APP_TITLE = "Simple Crosstab Application"
APP_VERSION = "v6.0"


def _default_upload_dir() -> str:
    return os.path.join(tempfile.gettempdir(), "crosstab_uploads")


UPLOAD_DIR = os.environ.get("CROSSTAB_UPLOAD_DIR", _default_upload_dir())
os.makedirs(UPLOAD_DIR, exist_ok=True)

app = FastAPI(title=APP_TITLE)
BASE_DIR = os.path.dirname(os.path.dirname(__file__))  # project root (Crosstab)
TEMPLATES_DIR = os.path.join(BASE_DIR, "templates")
templates = Jinja2Templates(directory=TEMPLATES_DIR)


STATIC_DIR = os.path.join(BASE_DIR, "static")
if os.path.isdir(STATIC_DIR):
    app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")


@app.get("/favicon.ico", include_in_schema=False)
def favicon():
    return RedirectResponse(url="/static/favicon.svg", status_code=307)


JOBS: Dict[str, Dict[str, Any]] = {}

YES_WORDS = {"yes", "y", "true", "t"}
NO_WORDS = {"no", "n", "false", "f"}


@dataclass(frozen=True)
class OrderedCategory:
    raw_code: str
    label: str
    order: int


@dataclass
class QuestionMapping:
    question: str
    categories: List[OrderedCategory]
    by_code: Dict[str, OrderedCategory]


@dataclass
class MappingBundle:
    questions: Dict[str, QuestionMapping]
    value_path: str
    text_path: str
    assumptions: List[str]
    source_kind: str = "paired"


@dataclass(frozen=True)
class MRGroup:
    stem: str
    columns: List[str]
    options: List[str]


@dataclass(frozen=True)
class ConceptColumn:
    family_key: str
    group_label: str
    concept_code: str
    concept_label: str
    concept_order: int
    source_column: str


CODEBOOK_REQUIRED_COLUMNS = {"question", "code", "label", "order"}


def _is_selected(expr: pl.Expr) -> pl.Expr:
    return (
        (expr.cast(pl.Float64, strict=False) == 1)
        | (
            expr.cast(pl.Utf8, strict=False)
            .str.to_lowercase()
            .is_in(["1", "true", "yes", "y", "t"])
        )
    )


def _is_dichotomy_series(s: Any) -> bool:
    try:
        if hasattr(s, "drop_nulls"):
            nonnull = s.drop_nulls()
            nonnull_len = nonnull.len()
            values = nonnull.to_list()[:5000]
            dtype = getattr(nonnull, "dtype", None)
        else:
            nonnull = s.dropna()
            nonnull_len = len(nonnull)
            values = nonnull.tolist()[:5000]
            dtype = getattr(nonnull, "dtype", None)
        if nonnull_len == 0:
            return False
        if dtype in (
            pl.Int64, pl.Int32, pl.Int16, pl.Int8,
            pl.UInt64, pl.UInt32, pl.UInt16, pl.UInt8,
            pl.Float64, pl.Float32
        ):
            vals = set(v for v in values if v is not None)
            return vals.issubset({0, 1, 0.0, 1.0})
        vals = set(str(v).strip().lower() for v in values)
        return vals.issubset({"0", "1"} | YES_WORDS | NO_WORDS)
    except Exception:
        return False


def _detect_dichotomy_columns(df: pl.DataFrame) -> List[str]:
    out: List[str] = []
    for c in df.columns:
        if _is_dichotomy_series(df[c]):
            out.append(c)
    return out


def _normalize_label_text(label: Any) -> str:
    text = "" if label is None else str(label)
    text = re.sub(r"\s+", " ", text.strip())
    text = re.sub(r"\s*([:：|])\s*", r" \1 ", text)
    text = re.sub(r"\s*-\s*", " - ", text)
    text = re.sub(r"\s*。\s*", "。", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _trim_stem_text(text: str) -> str:
    trimmed = re.sub(r"[\s:：|。-]+$", "", text.strip())
    return re.sub(r"\s+", " ", trimmed).strip()


def _split_label_stem_option(label: str) -> Optional[Tuple[str, str]]:
    normalized = _normalize_label_text(label)
    if not normalized:
        return None

    separators = ["\t", " | ", "：", ":", " - ", "。"]
    candidates: List[Tuple[str, str]] = []
    for separator in separators:
        if separator not in normalized:
            continue
        stem_part, option_part = normalized.rsplit(separator, 1)
        stem = _trim_stem_text(stem_part)
        option = option_part.strip()
        if not stem or not option:
            continue
        if len(stem) < 6:
            continue
        if stem == normalized or option == normalized:
            continue
        candidates.append((stem, option))

    if not candidates:
        return None
    candidates.sort(key=lambda item: (len(item[0]), -len(item[1])), reverse=True)
    return candidates[0]


def _longest_common_prefix(text1: str, text2: str) -> str:
    prefix_chars: List[str] = []
    for ch1, ch2 in zip(text1, text2):
        if ch1 != ch2:
            break
        prefix_chars.append(ch1)
    return "".join(prefix_chars)


def _build_mr_group(stem: str, columns: List[str], options: List[str]) -> MRGroup:
    return MRGroup(stem=stem, columns=list(columns), options=list(options))


def _group_from_split_candidates(df: Any, binary_columns: List[str]) -> Dict[str, MRGroup]:
    grouped: Dict[str, List[Tuple[str, str, bool]]] = {}
    binary_set = set(binary_columns)
    for column in list(df.columns):
        split = _split_label_stem_option(column)
        if split is None:
            continue
        stem, option = split
        grouped.setdefault(stem, []).append((column, option, column in binary_set))

    mr_groups: Dict[str, MRGroup] = {}
    for stem, items in grouped.items():
        if len(items) < 2:
            continue
        if not all(is_binary for _, _, is_binary in items):
            continue
        columns = [column for column, _, _ in items]
        options = [option for _, option, _ in items]
        if len({option.lower() for option in options}) != len(options):
            continue
        mr_groups[stem] = _build_mr_group(stem, columns, options)
    return mr_groups


def _group_from_similarity_fallback(df: pl.DataFrame, binary_columns: List[str], existing_stems: set[str]) -> Dict[str, MRGroup]:
    candidates: Dict[str, List[str]] = {}
    all_columns = list(df.columns)
    normalized_map = {column: _normalize_label_text(column) for column in all_columns}
    binary_set = set(binary_columns)

    for i, column_a in enumerate(binary_columns):
        text_a = normalized_map[column_a]
        for column_b in binary_columns[i + 1:]:
            text_b = normalized_map[column_b]
            ratio = SequenceMatcher(None, text_a, text_b).ratio()
            if ratio < 0.72:
                continue
            prefix = _trim_stem_text(_longest_common_prefix(text_a, text_b))
            if len(prefix) < 10 or prefix in existing_stems:
                continue
            if prefix == text_a or prefix == text_b:
                continue
            candidates.setdefault(prefix, [])
            if column_a not in candidates[prefix]:
                candidates[prefix].append(column_a)
            if column_b not in candidates[prefix]:
                candidates[prefix].append(column_b)

    mr_groups: Dict[str, MRGroup] = {}
    for stem, columns in candidates.items():
        matched_columns = [column for column in all_columns if normalized_map[column].startswith(stem)]
        if len(matched_columns) < 2:
            continue
        if not all(column in binary_set for column in matched_columns):
            continue
        options = []
        valid = True
        for column in matched_columns:
            option = normalized_map[column][len(stem):].strip(" :：|-。")
            if not option:
                valid = False
                break
            options.append(option)
        if not valid or len({option.lower() for option in options}) != len(options):
            continue
        mr_groups[stem] = _build_mr_group(stem, matched_columns, options)
    return mr_groups


def detect_mr_groups(df: Any) -> Dict[str, Any]:
    columns = list(df.columns)
    column_positions = {column: index for index, column in enumerate(columns)}
    binary_columns = [column for column in columns if _is_dichotomy_series(df[column])]
    exact_groups = _group_from_split_candidates(df, binary_columns)
    fallback_groups = _group_from_similarity_fallback(df, binary_columns, set(exact_groups.keys()))

    merged_groups = {**exact_groups, **fallback_groups}
    ordered_groups = sorted(
        merged_groups.values(),
        key=lambda group: min(column_positions[column] for column in group.columns),
    )
    grouped_columns = {column for group in ordered_groups for column in group.columns}

    return {
        "mr_groups": [
            {
                "question": group.stem,
                "stem": group.stem,
                "columns": list(group.columns),
                "options": list(group.options),
            }
            for group in ordered_groups
        ],
        "non_mr_columns": [column for column in columns if column not in grouped_columns],
    }


def _normalize_concept_group_label(label: str) -> str:
    text = _normalize_label_text(label)
    text = re.sub(r"^concept\s*\d+\s*:\s*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\b([A-Za-z]+\d+)_\d+\b", r"\1", text)
    text = re.sub(r"[_\-\s]+(?:c|concept)?\d+$", "", text, flags=re.IGNORECASE)
    return _normalize_label_text(text)


def _extract_concept_column(column: str) -> Optional[ConceptColumn]:
    text = _normalize_label_text(column)
    text = re.sub(r"^\d+\s*:\s*", "", text)

    prefix_match = re.match(r"^(concept)\s*(\d+)\s*:\s*(.+)$", text, flags=re.IGNORECASE)
    if prefix_match:
        concept_num = int(prefix_match.group(2))
        remainder = prefix_match.group(3).strip()
        group_label = _normalize_concept_group_label(remainder)
        family_key = _normalize_match_key(group_label)
        return ConceptColumn(
            family_key=family_key,
            group_label=group_label,
            concept_code=str(concept_num),
            concept_label=f"Concept {concept_num}",
            concept_order=concept_num,
            source_column=column,
        )

    suffix_match = re.match(r"^(.*?)(?:[_\-\s]+(?:c|concept)?(\d+))$", text, flags=re.IGNORECASE)
    if suffix_match:
        base = suffix_match.group(1).strip()
        concept_num = int(suffix_match.group(2))
        group_label = _normalize_concept_group_label(base)
        family_key = _normalize_match_key(group_label)
        return ConceptColumn(
            family_key=family_key,
            group_label=group_label,
            concept_code=str(concept_num),
            concept_label=f"Concept {concept_num}",
            concept_order=concept_num,
            source_column=column,
        )

    inline_match = re.search(r"\bconcept\s*(\d+)\b", text, flags=re.IGNORECASE)
    if inline_match:
        concept_num = int(inline_match.group(1))
        group_label = _normalize_concept_group_label(text)
        family_key = _normalize_match_key(group_label)
        if family_key and group_label != text:
            return ConceptColumn(
                family_key=family_key,
                group_label=group_label,
                concept_code=str(concept_num),
                concept_label=f"Concept {concept_num}",
                concept_order=concept_num,
                source_column=column,
            )

    return None


def detect_concept_groups(columns: List[str]) -> Dict[str, Any]:
    grouped: Dict[str, List[ConceptColumn]] = {}
    for column in columns:
        match = _extract_concept_column(str(column))
        if match is None:
            continue
        grouped.setdefault(match.family_key, []).append(match)

    concept_groups: List[Dict[str, Any]] = []
    member_to_group: Dict[str, Dict[str, Any]] = {}
    for family_key, members in grouped.items():
        distinct_concepts = {member.concept_code for member in members}
        if len(members) < 2 or len(distinct_concepts) < 2:
            continue
        ordered_members = sorted(members, key=lambda member: (member.concept_order, member.source_column))
        group = {
            "family_key": family_key,
            "label": ordered_members[0].group_label,
            "members": [
                {
                    "column": member.source_column,
                    "concept_code": member.concept_code,
                    "concept_label": member.concept_label,
                    "concept_order": member.concept_order,
                }
                for member in ordered_members
            ],
        }
        concept_groups.append(group)
        for member in group["members"]:
            member_to_group[str(member["column"])] = group

    concept_groups.sort(
        key=lambda group: min(int(member["concept_order"]) for member in group["members"])
    )
    return {
        "concept_groups": concept_groups,
        "member_to_group": member_to_group,
    }


def _mr_display_metadata(mr_detection: Optional[Dict[str, Any]], mr_cols: List[str]) -> Tuple[str, Dict[str, str]]:
    if not mr_detection:
        return "MR options", {column: column for column in mr_cols}

    display_labels: Dict[str, str] = {column: column for column in mr_cols}
    matched_stems: List[str] = []
    for group in mr_detection.get("mr_groups", []):
        stem = str(group.get("stem", "")).strip()
        columns = list(group.get("columns", []))
        options = list(group.get("options", []))
        option_map = {column: options[index] for index, column in enumerate(columns) if index < len(options)}
        group_has_selected = False
        for column in mr_cols:
            if column in option_map:
                display_labels[column] = option_map[column]
                group_has_selected = True
        if group_has_selected and stem:
            matched_stems.append(stem)

    unique_stems = list(dict.fromkeys(matched_stems))
    if len(unique_stems) == 1:
        return unique_stems[0], display_labels
    return "MR options", display_labels


def _pick_id_col(df: pl.DataFrame) -> Tuple[pl.DataFrame, str]:
    preferred = ["respid", "RESPID", "respondent_id", "RespondentID", "id", "ID"]
    for c in preferred:
        if c in df.columns:
            return df, c
    if "__rowid__" not in df.columns:
        df = df.with_row_index("__rowid__")
    return df, "__rowid__"


def _read_dataframe(path: str) -> pl.DataFrame:
    low = path.lower()
    if low.endswith(".csv"):
        return pl.read_csv(path, infer_schema_length=1000)
    if low.endswith(".xlsx"):
        if pd is None:
            raise RuntimeError("Excel upload requires pandas + openpyxl")
        pdf = pd.read_excel(path, engine="openpyxl")  # type: ignore
        return pl.from_pandas(pdf)
    raise RuntimeError("Unsupported file type (use .csv or .xlsx)")


def _read_excel_object_dataframe(path: str) -> "pd.DataFrame":
    if pd is None:
        raise RuntimeError("Excel mapping requires pandas + openpyxl")
    return pd.read_excel(path, engine="openpyxl", dtype=object)  # type: ignore[no-any-return]


def _normalize_response_value(value: Any) -> Optional[str]:
    if value is None:
        return None
    if pd is not None and pd.isna(value):
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        try:
            numeric = Decimal(stripped)
        except InvalidOperation:
            return stripped
        if numeric == numeric.to_integral_value():
            return str(int(numeric))
        return format(numeric.normalize(), "f")
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(Decimal(str(value)).normalize(), "f")
    text = str(value).strip()
    return text or None


def _normalize_display_label(value: Any, fallback: str) -> str:
    normalized = _normalize_response_value(value)
    return fallback if normalized is None else normalized


def normalize_values(value: Any) -> Optional[str]:
    return _normalize_response_value(value)


def _normalize_codebook_header(value: Any) -> str:
    text = _normalize_label_text(value).lower()
    text = re.sub(r"[^a-z0-9]+", "", text)
    aliases = {
        "question": "question",
        "questiontext": "question",
        "questionlabel": "question",
        "variable": "question",
        "variablename": "question",
        "qname": "question",
        "code": "code",
        "value": "code",
        "rawcode": "code",
        "responsecode": "code",
        "label": "label",
        "displaylabel": "label",
        "responselabel": "label",
        "text": "label",
        "order": "order",
        "displayorder": "order",
        "sortorder": "order",
        "sequence": "order",
    }
    return aliases.get(text, text)


def _read_excel_sheets_object_frames(path: str) -> Dict[str, "pd.DataFrame"]:
    if pd is None:
        raise RuntimeError("Excel mapping requires pandas + openpyxl")
    with pd.ExcelFile(path, engine="openpyxl") as workbook:
        return {sheet: workbook.parse(sheet, dtype=object) for sheet in workbook.sheet_names}


def _extract_codebook_frame(value_path: str) -> Optional["pd.DataFrame"]:
    for _, frame in _read_excel_sheets_object_frames(value_path).items():
        renamed = {column: _normalize_codebook_header(column) for column in frame.columns}
        canonical = frame.rename(columns=renamed)
        if CODEBOOK_REQUIRED_COLUMNS.issubset(set(canonical.columns)):
            return canonical.loc[:, ["question", "code", "label", "order"]].copy()
    return None


def _build_question_mappings_from_codebook(codebook_df: "pd.DataFrame") -> MappingBundle:
    cleaned = codebook_df.copy()
    for column in ["question", "code", "label", "order"]:
        cleaned[column] = cleaned[column].map(lambda value: None if (pd is not None and pd.isna(value)) else value)

    if bool(cleaned["question"].isna().any()):
        raise RuntimeError("Value.xlsx contains mapping rows with missing question values.")
    if bool(cleaned["code"].isna().any()):
        raise RuntimeError("Value.xlsx contains mapping rows with missing code values.")
    if bool(cleaned["label"].isna().any()):
        raise RuntimeError("Value.xlsx contains mapping rows with missing label values.")
    if bool(cleaned["order"].isna().any()):
        raise RuntimeError("Value.xlsx contains mapping rows with missing order values.")

    questions: Dict[str, QuestionMapping] = {}
    for raw_question, frame in cleaned.groupby("question", dropna=False, sort=False):
        question = _normalize_label_text(raw_question)
        categories: List[OrderedCategory] = []
        by_code: Dict[str, OrderedCategory] = {}
        seen_orders: Dict[int, str] = {}
        for _, row in frame.iterrows():
            code = _normalize_response_value(row["code"])
            if code is None:
                raise RuntimeError(f"Question '{question}' contains an empty code in Value.xlsx.")
            label = _normalize_display_label(row["label"], fallback=code)
            order_text = _normalize_response_value(row["order"])
            if order_text is None:
                raise RuntimeError(f"Question '{question}' contains an empty order in Value.xlsx.")
            try:
                order = int(Decimal(order_text))
            except (InvalidOperation, ValueError):
                raise RuntimeError(
                    f"Question '{question}' contains non-numeric order '{row['order']}' in Value.xlsx."
                )
            if code in by_code:
                raise RuntimeError(f"Question '{question}' has duplicate code '{code}' in Value.xlsx.")
            if order in seen_orders:
                raise RuntimeError(
                    f"Question '{question}' has duplicate order '{order}' for codes "
                    f"'{seen_orders[order]}' and '{code}' in Value.xlsx."
                )
            category = OrderedCategory(raw_code=code, label=label, order=order)
            categories.append(category)
            by_code[code] = category
            seen_orders[order] = code
        categories.sort(key=lambda category: category.order)
        questions[question] = QuestionMapping(question=question, categories=categories, by_code=by_code)

    return MappingBundle(
        questions=questions,
        value_path="",
        text_path="",
        assumptions=["Value.xlsx supplies question-specific code, label, and explicit display order."],
        source_kind="codebook",
    )


def _sort_row_categories(categories: List[OrderedCategory], row_sort_mode: str) -> List[OrderedCategory]:
    numeric_rows: List[Tuple[Decimal, OrderedCategory]] = []
    for category in categories:
        try:
            numeric_rows.append((Decimal(category.raw_code), category))
        except InvalidOperation:
            return list(categories)

    reverse = row_sort_mode == "code_desc"
    numeric_rows.sort(key=lambda item: item[0], reverse=reverse)
    return [category for _, category in numeric_rows]


def _sort_column_categories(categories: List[OrderedCategory]) -> List[OrderedCategory]:
    numeric_columns: List[Tuple[Decimal, OrderedCategory]] = []
    for category in categories:
        try:
            numeric_columns.append((Decimal(category.raw_code), category))
        except InvalidOperation:
            return list(categories)
    numeric_columns.sort(key=lambda item: item[0])
    return [category for _, category in numeric_columns]


def _apply_column_ordering(bundle: MappingBundle, categories: List[OrderedCategory]) -> List[OrderedCategory]:
    if bundle.source_kind == "codebook":
        return sorted(categories, key=lambda category: category.order)
    return _sort_column_categories(categories)


def _build_header_layout(columns: List[str]) -> Dict[str, Any]:
    if not any(" | " in str(column) for column in columns):
        return {
            "two_layer": False,
            "top_groups": [],
            "bottom_labels": [str(column) for column in columns],
            "group_start_indices": [],
        }

    top_labels: List[str] = []
    bottom_labels: List[str] = []
    for column in columns:
        text = str(column)
        if " | " in text:
            parent, child = text.split(" | ", 1)
            top_labels.append(parent.strip())
            bottom_labels.append(child.strip())
        else:
            top_labels.append(text.strip())
            bottom_labels.append("")

    groups: List[Dict[str, Any]] = []
    group_start_indices: List[int] = []
    for index, label in enumerate(top_labels):
        if groups and groups[-1]["label"] == label:
            groups[-1]["span"] += 1
        else:
            groups.append({"label": label, "span": 1})
            if index > 0:
                group_start_indices.append(index)

    return {
        "two_layer": True,
        "top_groups": groups,
        "bottom_labels": bottom_labels,
        "group_start_indices": group_start_indices,
    }


def _build_box_row(
    label: str,
    categories: List[OrderedCategory],
    count_lookup: Dict[Tuple[str, str], float],
    col_categories: List[OrderedCategory],
    row_totals: Dict[str, float],
    col_totals: Dict[str, float],
    grand_total: float,
    kind: str,
    include_totals: bool,
) -> Dict[str, Any]:
    row: Dict[str, Any] = {"__label__": label, "__is_box__": True}
    combined_row_total = sum(row_totals.get(cat.raw_code, 0) for cat in categories)
    for col_cat in col_categories:
        count = sum(count_lookup.get((row_cat.raw_code, col_cat.raw_code), 0) for row_cat in categories)
        if kind == "counts":
            row[col_cat.label] = count
        elif kind == "rowpct":
            row[col_cat.label] = (count / combined_row_total * 100.0) if combined_row_total > 0 else 0.0
        else:
            col_total = col_totals.get(col_cat.raw_code, 0)
            row[col_cat.label] = (count / col_total * 100.0) if col_total > 0 else 0.0
    if include_totals:
        if kind == "counts":
            row["Total"] = combined_row_total
        elif kind == "rowpct":
            row["Total"] = 100.0 if combined_row_total > 0 else 0.0
        else:
            row["Total"] = (combined_row_total / grand_total * 100.0) if grand_total > 0 else 0.0
    return row


def _build_custom_group_row(
    label: str,
    selected_codes: List[str],
    row_categories: List[OrderedCategory],
    count_lookup: Dict[Tuple[str, str], float],
    col_categories: List[OrderedCategory],
    row_totals: Dict[str, float],
    col_totals: Dict[str, float],
    grand_total: float,
    kind: str,
    include_totals: bool,
) -> Optional[Dict[str, Any]]:
    normalized_label = str(label).strip()
    wanted_codes = {str(code).strip() for code in selected_codes if str(code).strip()}
    if not normalized_label or not wanted_codes:
        return None
    categories = [category for category in row_categories if category.raw_code in wanted_codes]
    if not categories:
        return None
    return _build_box_row(
        label=normalized_label,
        categories=categories,
        count_lookup=count_lookup,
        col_categories=col_categories,
        row_totals=row_totals,
        col_totals=col_totals,
        grand_total=grand_total,
        kind=kind,
        include_totals=include_totals,
    )


def _build_custom_group_row_from_counts(
    label: str,
    group_counts: Dict[str, float],
    col_categories: List[OrderedCategory],
    col_denominators: Dict[str, float],
    grand_total: float,
    kind: str,
    include_totals: bool,
) -> Optional[Dict[str, Any]]:
    normalized_label = str(label).strip()
    if not normalized_label:
        return None
    row: Dict[str, Any] = {"__label__": normalized_label, "__is_box__": True}
    row_total = sum(float(group_counts.get(str(col.raw_code), 0)) for col in col_categories)
    for col_cat in col_categories:
        count = float(group_counts.get(str(col_cat.raw_code), 0))
        if kind == "counts":
            row[col_cat.label] = count
        elif kind == "rowpct":
            row[col_cat.label] = (count / row_total * 100.0) if row_total > 0 else 0.0
        else:
            denominator = float(col_denominators.get(str(col_cat.raw_code), 0))
            row[col_cat.label] = (count / denominator * 100.0) if denominator > 0 else 0.0
    if include_totals:
        if kind == "counts":
            row["Total"] = row_total
        elif kind == "rowpct":
            row["Total"] = 100.0 if row_total > 0 else 0.0
        else:
            row["Total"] = (row_total / grand_total * 100.0) if grand_total > 0 else 0.0
    return row


def _parse_custom_group_definitions(
    definitions_text: str,
    single_label: str = "",
    single_codes: Optional[List[str]] = None,
) -> List[Dict[str, Any]]:
    groups: List[Dict[str, Any]] = []
    lines = [str(line).strip() for line in str(definitions_text or "").splitlines()]
    for line in lines:
        if not line:
            continue
        if "=" not in line:
            raise RuntimeError(
                "Custom groups must use the format 'Label = code1, code2'."
            )
        label, codes_text = line.split("=", 1)
        parsed_codes = [code.strip() for code in re.split(r"[,|]", codes_text) if code.strip()]
        if not label.strip() or not parsed_codes:
            raise RuntimeError(
                "Custom groups must use the format 'Label = code1, code2'."
            )
        groups.append({"label": label.strip(), "codes": parsed_codes})

    fallback_codes = [str(code).strip() for code in list(single_codes or []) if str(code).strip()]
    if str(single_label).strip() and fallback_codes:
        groups.append({"label": str(single_label).strip(), "codes": fallback_codes})
    return groups


def _groupable_question_options(bundle: MappingBundle) -> Dict[str, List[Dict[str, Any]]]:
    return {
        question: [
            {"code": category.raw_code, "label": category.label, "order": category.order}
            for category in mapping.categories
        ]
        for question, mapping in bundle.questions.items()
    }


def _short_column_label(column_name: str) -> str:
    value = str(column_name or "").strip()
    match = re.match(r"^\s*\d+\s*:\s*(.+)$", value)
    if match:
        return match.group(1).strip()
    return value


def _column_display_labels(columns: List[str]) -> Dict[str, str]:
    labels: Dict[str, str] = {}
    used: Dict[str, int] = {}
    for column in columns:
        base_label = _short_column_label(column) or str(column)
        count = used.get(base_label, 0)
        used[base_label] = count + 1
        labels[str(column)] = base_label if count == 0 else f"{base_label} ({count + 1})"
    return labels


def _build_filter_value_options(
    df: "pd.DataFrame",
    bundle: Optional["MappingBundle"],
) -> Dict[str, List[Dict[str, str]]]:
    options_by_column: Dict[str, List[Dict[str, str]]] = {}
    bundle_questions = getattr(bundle, "questions", {}) if bundle is not None else {}
    for column in [str(col) for col in df.columns]:
        if column in bundle_questions:
            try:
                mapping = bundle_questions[column]
                options_by_column[column] = [
                    {
                        "value": str(category.raw_code),
                        "label": f"{category.raw_code} - {category.label}",
                    }
                    for category in mapping.categories
                ]
                continue
            except Exception:
                pass

        try:
            if pd is not None:
                unique_values = df[column].dropna().unique()
                options_by_column[column] = [
                    {"value": str(val), "label": str(val)}
                    for val in unique_values[:50]
                    if val is not None
                ]
            else:
                unique_values = df[column].drop_nulls().unique()
                options_by_column[column] = [
                    {"value": str(val), "label": str(val)}
                    for val in unique_values[:50]
                ]
        except Exception:
            options_by_column[column] = []
    return options_by_column


def _validate_mapping_frames(text_df: "pd.DataFrame", value_df: "pd.DataFrame") -> None:
    text_columns = [str(c) for c in text_df.columns]
    value_columns = [str(c) for c in value_df.columns]
    if text_columns != value_columns:
        raise RuntimeError("Uploaded text workbook and Value.xlsx must have identical columns.")
    if len(text_df.index) != len(value_df.index):
        raise RuntimeError("Uploaded text workbook and Value.xlsx must have identical row counts.")


def _build_question_mappings(text_df: "pd.DataFrame", value_df: "pd.DataFrame") -> MappingBundle:
    _validate_mapping_frames(text_df, value_df)
    questions: Dict[str, QuestionMapping] = {}
    assumptions = [
        "Value.xlsx stores respondent-level raw codes rather than a standalone codebook.",
        "Display labels are paired from the uploaded text workbook by matching row/column positions.",
        "Category order is assigned by first appearance of each raw code while scanning Value.xlsx row order.",
    ]

    for column in text_df.columns:
        question = str(column)
        categories: List[OrderedCategory] = []
        by_code: Dict[str, OrderedCategory] = {}
        for raw_value, text_value in zip(value_df[column].tolist(), text_df[column].tolist()):
            code = _normalize_response_value(raw_value)
            if code is None:
                continue
            label = _normalize_display_label(text_value, fallback=code)
            existing = by_code.get(code)
            if existing is None:
                category = OrderedCategory(raw_code=code, label=label, order=len(categories))
                categories.append(category)
                by_code[code] = category
            elif existing.label != label:
                raise RuntimeError(
                    f"Conflicting labels for question '{question}' and code '{code}': "
                    f"'{existing.label}' vs '{label}'"
                )
        questions[question] = QuestionMapping(question=question, categories=categories, by_code=by_code)

    return MappingBundle(
        questions=questions,
        value_path="",
        text_path="",
        assumptions=assumptions,
        source_kind="paired",
    )


def _load_mapping_bundle(text_path: str, value_path: str) -> MappingBundle:
    codebook_df = _extract_codebook_frame(value_path)
    if codebook_df is not None:
        bundle = _build_question_mappings_from_codebook(codebook_df)
    else:
        text_df = _read_excel_object_dataframe(text_path)
        value_df = _read_excel_object_dataframe(value_path)
        bundle = _build_question_mappings(text_df, value_df)
    bundle.value_path = value_path
    bundle.text_path = text_path
    return bundle


def load_mapping(text_path: str, value_path: str) -> MappingBundle:
    return _load_mapping_bundle(text_path, value_path)


def _looks_like_value_workbook(filename: str) -> bool:
    name = os.path.basename(filename).strip().lower()
    return name == "value.xlsx"


def _resolve_uploaded_workbook_roles(
    file_path_a: str,
    file_name_a: str,
    file_path_b: str,
    file_name_b: str,
) -> Tuple[str, str, MappingBundle]:
    candidates: List[Tuple[str, str]] = []
    a_is_value = _looks_like_value_workbook(file_name_a)
    b_is_value = _looks_like_value_workbook(file_name_b)

    if a_is_value and not b_is_value:
        candidates.append((file_path_b, file_path_a))
    elif b_is_value and not a_is_value:
        candidates.append((file_path_a, file_path_b))
    else:
        candidates.append((file_path_a, file_path_b))
        candidates.append((file_path_b, file_path_a))

    seen: set[Tuple[str, str]] = set()
    errors: List[str] = []
    for text_path, value_path in candidates:
        key = (text_path, value_path)
        if key in seen:
            continue
        seen.add(key)
        try:
            bundle = _load_mapping_bundle(text_path, value_path)
            return text_path, value_path, bundle
        except Exception as e:
            errors.append(f"text={os.path.basename(text_path)}, value={os.path.basename(value_path)} -> {e}")

    joined = " | ".join(errors) if errors else "no compatible workbook pairing found"
    raise RuntimeError(f"Failed to identify text workbook and Value.xlsx automatically: {joined}")


def _question_mapping_or_error(bundle: MappingBundle, question: str) -> QuestionMapping:
    mapping = bundle.questions.get(question)
    if mapping is None:
        raise RuntimeError(f"Question '{question}' exists in dataset but has no mapping in Value.xlsx.")
    if not mapping.categories:
        raise RuntimeError(f"Question '{question}' has no mapped categories in Value.xlsx.")
    return mapping


def _validate_question_codes(question: str, mapping: QuestionMapping, codes: List[Optional[str]]) -> None:
    missing = sorted({code for code in codes if code is not None and code not in mapping.by_code})
    if missing:
        shown = ", ".join(missing[:10])
        suffix = "" if len(missing) <= 10 else f" (+{len(missing) - 10} more)"
        raise RuntimeError(
            f"Question '{question}' contains raw code(s) missing from Value.xlsx: {shown}{suffix}"
        )


def apply_ordering(categories: List[OrderedCategory], row_sort_mode: str = "value_order") -> List[OrderedCategory]:
    if row_sort_mode == "value_order":
        return sorted(categories, key=lambda category: category.order)
    return _sort_row_categories(categories, row_sort_mode)


def _normalize_match_key(value: Any) -> str:
    return _normalize_label_text(value).casefold()


def _shared_mr_question(mr_detection: Optional[Dict[str, Any]], mr_cols: List[str]) -> Optional[str]:
    if not mr_detection:
        return None
    matched_questions: List[str] = []
    selected = set(mr_cols)
    for group in mr_detection.get("mr_groups", []):
        group_columns = set(group.get("columns", []))
        if selected & group_columns:
            question = str(group.get("question") or group.get("stem") or "").strip()
            if question:
                matched_questions.append(question)
    unique = list(dict.fromkeys(matched_questions))
    return unique[0] if len(unique) == 1 else None


def _build_mr_row_categories(
    bundle: MappingBundle,
    mr_cols: List[str],
    mr_detection: Optional[Dict[str, Any]],
) -> Tuple[str, List[OrderedCategory], Dict[str, OrderedCategory]]:
    mr_row_label, mr_display_labels = _mr_display_metadata(mr_detection, mr_cols)
    shared_question = _shared_mr_question(mr_detection, mr_cols)

    if shared_question and shared_question in bundle.questions:
        mapping = _question_mapping_or_error(bundle, shared_question)
        selected_by_option = {
            _normalize_match_key(mr_display_labels.get(column, column)): column
            for column in mr_cols
        }
        row_categories: List[OrderedCategory] = []
        column_to_category: Dict[str, OrderedCategory] = {}
        for category in mapping.categories:
            matched_column = selected_by_option.get(_normalize_match_key(category.label))
            if matched_column:
                row_categories.append(category)
                column_to_category[matched_column] = category
        missing_columns = [column for column in mr_cols if column not in column_to_category]
        if missing_columns:
            sample = ", ".join(missing_columns[:5])
            raise RuntimeError(
                f"MR option(s) for question '{shared_question}' are missing from Value.xlsx mapping: {sample}"
            )
        return shared_question, row_categories, column_to_category

    row_categories = [
        OrderedCategory(raw_code=question, label=mr_display_labels.get(question, question), order=index)
        for index, question in enumerate(mr_cols)
    ]
    column_to_category = {column: row_categories[index] for index, column in enumerate(mr_cols)}
    return mr_row_label, row_categories, column_to_category


def _mapped_series(
    df: "pd.DataFrame",
    bundle: MappingBundle,
    question: str,
    include_all_categories: bool,
) -> Tuple["pd.DataFrame", List[OrderedCategory]]:
    mapping = _question_mapping_or_error(bundle, question)
    codes = [_normalize_response_value(v) for v in df[question].tolist()]
    _validate_question_codes(question, mapping, codes)

    rows: List[Dict[str, Any]] = []
    for code in codes:
        if code is None:
            continue
        cat = mapping.by_code[code]
        rows.append({"code": cat.raw_code, "label": cat.label, "order": cat.order})

    frame = pd.DataFrame(rows, columns=["code", "label", "order"])  # type: ignore[union-attr]
    present_codes = {r["code"] for r in rows}
    categories = list(mapping.categories) if include_all_categories else [
        cat for cat in mapping.categories if cat.raw_code in present_codes
    ]
    return frame, categories


def _mapped_codes_for_question(
    df: "pd.DataFrame",
    bundle: MappingBundle,
    question: str,
) -> Tuple[List[Optional[str]], QuestionMapping]:
    mapping = _question_mapping_or_error(bundle, question)
    codes = [_normalize_response_value(v) for v in df[question].tolist()]
    _validate_question_codes(question, mapping, codes)
    return codes, mapping


def _composite_codes_for_questions(
    df: "pd.DataFrame",
    bundle: MappingBundle,
    question1: str,
    question2: str,
) -> List[Optional[str]]:
    codes1, _ = _mapped_codes_for_question(df, bundle, question1)
    codes2, _ = _mapped_codes_for_question(df, bundle, question2)
    composite_codes: List[Optional[str]] = []
    for code1, code2 in zip(codes1, codes2):
        if code1 is None or code2 is None:
            composite_codes.append(None)
        else:
            composite_codes.append(f"{code1} | {code2}")
    return composite_codes


def _mapped_composite_series(
    df: "pd.DataFrame",
    bundle: MappingBundle,
    question1: str,
    question2: str,
    include_all_categories: bool,
) -> Tuple["pd.DataFrame", List[OrderedCategory]]:
    mapping1 = _question_mapping_or_error(bundle, question1)
    mapping2 = _question_mapping_or_error(bundle, question2)

    rows: List[Dict[str, Any]] = []
    seen_codes: set[str] = set()
    for raw1, raw2 in zip(df[question1].tolist(), df[question2].tolist()):
        code1 = _normalize_response_value(raw1)
        code2 = _normalize_response_value(raw2)
        if code1 is None or code2 is None:
            continue
        if code1 not in mapping1.by_code:
            raise RuntimeError(f"Question '{question1}' contains raw code(s) missing from Value.xlsx: {code1}")
        if code2 not in mapping2.by_code:
            raise RuntimeError(f"Question '{question2}' contains raw code(s) missing from Value.xlsx: {code2}")
        cat1 = mapping1.by_code[code1]
        cat2 = mapping2.by_code[code2]
        composite_code = f"{cat1.raw_code} | {cat2.raw_code}"
        composite_label = f"{cat1.label} | {cat2.label}"
        composite_order = cat1.order * 10000 + cat2.order
        rows.append({"code": composite_code, "label": composite_label, "order": composite_order})
        seen_codes.add(composite_code)

    categories: List[OrderedCategory] = []
    for cat1 in mapping1.categories:
        for cat2 in mapping2.categories:
            composite_code = f"{cat1.raw_code} | {cat2.raw_code}"
            if include_all_categories or composite_code in seen_codes:
                categories.append(
                    OrderedCategory(
                        raw_code=composite_code,
                        label=f"{cat1.label} | {cat2.label}",
                        order=cat1.order * 10000 + cat2.order,
                    )
                )
    frame = pd.DataFrame(rows, columns=["code", "label", "order"])  # type: ignore[union-attr]
    return frame, categories


def _resolve_weight_values(
    df: "pd.DataFrame",
    weight_col: Optional[str],
) -> Tuple[List[float], Optional[str]]:
    if not weight_col or weight_col == "(none)":
        return [1.0] * len(df.index), None
    if weight_col not in df.columns:
        raise RuntimeError("Selected weight column not found in dataset.")
    raw_series = df[weight_col]
    numeric_series = pd.to_numeric(raw_series, errors="coerce")
    invalid_mask = raw_series.notna() & numeric_series.isna()
    if bool(invalid_mask.any()):
        raise RuntimeError(f"Weight column '{weight_col}' contains non-numeric values.")
    numeric_series = numeric_series.fillna(0.0)
    if bool((numeric_series < 0).any()):
        raise RuntimeError(f"Weight column '{weight_col}' contains negative values.")
    return [float(value) for value in numeric_series.tolist()], weight_col


def _prepare_column_dimension(
    df: "pd.DataFrame",
    bundle: MappingBundle,
    col_var: str,
    col_var2: Optional[str],
    include_all_categories: bool,
    total_length_override: Optional[int] = None,
) -> Tuple["pd.DataFrame", List[OrderedCategory], str]:
    if not col_var or col_var == "(none)":
        length = int(total_length_override) if total_length_override is not None else len(df.index)
        rows = [{"code": "__total__", "label": "Total", "order": 0} for _ in range(max(length, 0))]
        frame = pd.DataFrame(rows, columns=["code", "label", "order"])  # type: ignore[union-attr]
        return frame, [OrderedCategory(raw_code="__total__", label="Total", order=0)], "Total"
    if not col_var2 or col_var2 == "(none)":
        frame, categories = _mapped_series(df, bundle, col_var, include_all_categories=include_all_categories)
        return frame, _apply_column_ordering(bundle, categories), col_var
    frame, categories = _mapped_composite_series(
        df,
        bundle,
        col_var,
        col_var2,
        include_all_categories=include_all_categories,
    )
    return frame, _apply_column_ordering(bundle, categories), f"{col_var} | {col_var2}"


def _build_concept_row_categories(
    bundle: MappingBundle,
    concept_group: Dict[str, Any],
) -> Tuple[str, List[OrderedCategory], Dict[str, OrderedCategory]]:
    merged_by_code: Dict[str, OrderedCategory] = {}
    row_label = str(concept_group.get("label", "Concept comparison"))

    for member in concept_group.get("members", []):
        column = str(member.get("column", ""))
        mapping = _question_mapping_or_error(bundle, column)
        for category in mapping.categories:
            existing = merged_by_code.get(category.raw_code)
            if existing is None:
                merged_by_code[category.raw_code] = OrderedCategory(
                    raw_code=category.raw_code,
                    label=category.label,
                    order=category.order,
                )
            elif existing.label != category.label:
                raise RuntimeError(
                    f"Concept comparison cannot merge '{row_label}' because code "
                    f"'{category.raw_code}' has conflicting labels: '{existing.label}' vs '{category.label}'."
                )
            elif category.order < existing.order:
                merged_by_code[category.raw_code] = OrderedCategory(
                    raw_code=category.raw_code,
                    label=category.label,
                    order=category.order,
                )

    row_categories = sorted(merged_by_code.values(), key=lambda category: category.order)
    if not row_categories:
        raise RuntimeError(f"Concept comparison is unavailable for '{row_label}' because no mapped categories were found.")
    return row_label, row_categories, merged_by_code


def tabulate_sr_concept_comparison(
    df: "pd.DataFrame",
    bundle: MappingBundle,
    concept_group: Dict[str, Any],
    top_banner_var: Optional[str],
    include_totals: bool,
    include_base: bool,
    out_counts: bool,
    out_rowpct: bool,
    out_colpct: bool,
    show_row_codes: bool = False,
    row_sort_mode: str = "code_asc",
    include_top2_box: bool = False,
    include_bottom2_box: bool = False,
    custom_group_label: str = "",
    custom_group_codes: Optional[List[str]] = None,
    custom_groups: Optional[List[Dict[str, Any]]] = None,
    hide_grouped_codes: bool = False,
    weight_col: Optional[str] = None,
) -> Dict[str, Any]:
    row_label, row_categories, merged_by_code = _build_concept_row_categories(bundle, concept_group)
    weights, resolved_weight_col = _resolve_weight_values(df, weight_col)
    concept_categories = [
        OrderedCategory(
            raw_code=str(member.get("concept_code", "")),
            label=f"C{int(member.get('concept_order', 0))}",
            order=int(member.get("concept_order", 0)),
        )
        for member in concept_group.get("members", [])
    ]
    if top_banner_var:
        top_frame, top_categories = _mapped_series(df, bundle, top_banner_var, include_all_categories=True)
        top_categories = _apply_column_ordering(bundle, top_categories)
        top_codes, _ = _mapped_codes_for_question(df, bundle, top_banner_var)
        col_categories: List[OrderedCategory] = []
        for top_category in top_categories:
            for concept_category in concept_categories:
                col_categories.append(
                    OrderedCategory(
                        raw_code=f"{top_category.raw_code} | {concept_category.raw_code}",
                        label=f"{top_category.label} | {concept_category.label}",
                        order=top_category.order * 10000 + concept_category.order,
                    )
                )
    else:
        col_categories = concept_categories
        top_codes = []

    rows: List[Dict[str, Any]] = []
    for member in concept_group.get("members", []):
        column = str(member.get("column", ""))
        concept_code = str(member.get("concept_code", ""))
        mapping = _question_mapping_or_error(bundle, column)
        codes = [_normalize_response_value(v) for v in df[column].tolist()]
        _validate_question_codes(column, mapping, codes)
        for row_index, code in enumerate(codes):
            if code is None:
                continue
            weight = weights[row_index] if row_index < len(weights) else 0.0
            if weight == 0:
                continue
            if code not in merged_by_code:
                raise RuntimeError(
                    f"Concept comparison cannot map code '{code}' from question '{column}'."
                )
            if top_banner_var:
                top_code = top_codes[row_index] if row_index < len(top_codes) else None
                if top_code is None:
                    continue
                col_code = f"{top_code} | {concept_code}"
            else:
                col_code = concept_code
            rows.append({"row_code": code, "col_code": col_code, "count": weight})

    counts = pd.DataFrame(rows, columns=["row_code", "col_code", "count"])  # type: ignore[union-attr]
    if not counts.empty:
        counts = counts.groupby(["row_code", "col_code"], dropna=False)["count"].sum().reset_index()
    tables = _format_ordered_tables(
        counts,
        row_categories,
        col_categories,
        include_totals=include_totals,
        include_base=include_base,
        out_counts=out_counts,
        out_rowpct=out_rowpct,
        out_colpct=out_colpct,
        show_row_codes=show_row_codes,
        row_sort_mode="value_order" if bundle.source_kind == "codebook" else row_sort_mode,
        include_top2_box=include_top2_box,
        include_bottom2_box=include_bottom2_box,
        custom_group_label=custom_group_label,
        custom_group_codes=custom_group_codes,
        custom_groups=custom_groups,
        hide_grouped_codes=hide_grouped_codes,
    )
    return {
        "mode": "sr",
        "row_label": row_label,
        "row_label_display": _preview_axis_label(row_label, "Responses"),
        "col_label": f"{top_banner_var} | Concept" if top_banner_var else "Concept",
        "col_label_display": _preview_axis_label(top_banner_var, "Concept")
        if top_banner_var
        else "Concept",
        "tables": tables,
        "validation": [],
        "comparison_mode": "concept",
        "weight_col": resolved_weight_col,
    }


def _format_ordered_tables(
    counts_df: "pd.DataFrame",
    row_categories: List[OrderedCategory],
    col_categories: List[OrderedCategory],
    include_totals: bool,
    include_base: bool,
    out_counts: bool,
    out_rowpct: bool,
    out_colpct: bool,
    show_row_codes: bool,
    row_sort_mode: str,
    include_top2_box: bool = False,
    include_bottom2_box: bool = False,
    custom_group_label: str = "",
    custom_group_codes: Optional[List[str]] = None,
    custom_groups: Optional[List[Dict[str, Any]]] = None,
    hide_grouped_codes: bool = False,
    base_col_totals_override: Optional[Dict[str, int]] = None,
    base_grand_total_override: Optional[int] = None,
    colpct_denominator_override: Optional[Dict[str, int]] = None,
    custom_group_count_overrides: Optional[Dict[str, Dict[str, int]]] = None,
) -> List[Dict[str, Any]]:
    row_categories = apply_ordering(row_categories, row_sort_mode)
    has_single_total_banner = len(col_categories) == 1 and str(col_categories[0].label) == "Total"
    include_grand_total_column = include_totals and not has_single_total_banner
    columns = [cat.label for cat in col_categories] + (["Total"] if include_grand_total_column else [])
    count_lookup: Dict[Tuple[str, str], float] = {}
    row_totals: Dict[str, float] = {cat.raw_code: 0.0 for cat in row_categories}
    col_totals: Dict[str, float] = {cat.raw_code: 0.0 for cat in col_categories}

    for _, rec in counts_df.iterrows():
        row_code = str(rec["row_code"])
        col_code = str(rec["col_code"])
        value = float(rec["count"])
        count_lookup[(row_code, col_code)] = value
        row_totals[row_code] = row_totals.get(row_code, 0) + value
        col_totals[col_code] = col_totals.get(col_code, 0) + value

    grand_total = sum(row_totals.values())
    base_col_totals = {str(key): float(value) for key, value in dict(base_col_totals_override or col_totals).items()}
    base_grand_total = float(base_grand_total_override) if base_grand_total_override is not None else grand_total
    colpct_denominators = {str(key): float(value) for key, value in dict(colpct_denominator_override or col_totals).items()}
    hidden_group_codes: set[str] = set()
    if hide_grouped_codes:
        configured_groups = list(custom_groups or [])
        if not configured_groups and (custom_group_label or custom_group_codes):
            configured_groups = [
                {"label": custom_group_label, "codes": list(custom_group_codes or [])}
            ]
        for group in configured_groups:
            hidden_group_codes.update(
                str(code).strip()
                for code in list(group.get("codes", []) or [])
                if str(code).strip()
            )
    visible_row_categories = [
        category for category in row_categories
        if category.raw_code not in hidden_group_codes
    ]

    def display_row_label(category: OrderedCategory) -> str:
        if not show_row_codes:
            return category.label
        return f"{category.raw_code} - {category.label}"

    def build(kind: str) -> Dict[str, Any]:
        rows_out: List[Dict[str, Any]] = []
        if include_base:
            base_row = {"__label__": "Base"}
            for cat in col_categories:
                base_row[cat.label] = base_col_totals.get(cat.raw_code, 0)
            if include_grand_total_column:
                base_row["Total"] = base_grand_total
            rows_out.append(base_row)

        top2_categories = row_categories[:2] if row_sort_mode == "code_desc" else row_categories[-2:]
        bottom2_categories = row_categories[-2:] if row_sort_mode == "code_desc" else row_categories[:2]
        pre_rows: List[Dict[str, Any]] = []
        post_rows: List[Dict[str, Any]] = []
        if include_top2_box and len(row_categories) >= 2:
            top2_row = _build_box_row(
                "Top 2 Box",
                top2_categories,
                count_lookup,
                col_categories,
                row_totals,
                col_totals,
                grand_total,
                kind,
                include_totals,
            )
            if row_sort_mode == "code_desc":
                pre_rows.append(top2_row)
            else:
                post_rows.append(top2_row)
        if include_bottom2_box and len(row_categories) >= 2:
            bottom2_row = _build_box_row(
                "Bottom 2 Box",
                bottom2_categories,
                count_lookup,
                col_categories,
                row_totals,
                col_totals,
                grand_total,
                kind,
                include_totals,
            )
            if row_sort_mode == "code_desc":
                post_rows.append(bottom2_row)
            else:
                pre_rows.append(bottom2_row)
        configured_groups = list(custom_groups or [])
        if not configured_groups and (custom_group_label or custom_group_codes):
            configured_groups = [
                {"label": custom_group_label, "codes": list(custom_group_codes or [])}
            ]
        for group in configured_groups:
            group_label = str(group.get("label", ""))
            group_override = (custom_group_count_overrides or {}).get(group_label)
            if group_override is not None:
                custom_group_row = _build_custom_group_row_from_counts(
                    label=group_label,
                    group_counts=group_override,
                    col_categories=col_categories,
                    col_denominators=colpct_denominators,
                    grand_total=base_grand_total,
                    kind=kind,
                    include_totals=include_grand_total_column,
                )
            else:
                custom_group_row = _build_custom_group_row(
                    label=group_label,
                    selected_codes=list(group.get("codes", []) or []),
                    row_categories=row_categories,
                    count_lookup=count_lookup,
                    col_categories=col_categories,
                    row_totals=row_totals,
                    col_totals=col_totals,
                    grand_total=grand_total,
                    kind=kind,
                    include_totals=include_grand_total_column,
                )
            if custom_group_row is not None:
                pre_rows.append(custom_group_row)

        rows_out.extend(pre_rows)

        for row_cat in visible_row_categories:
            row = {"__label__": display_row_label(row_cat), "__code__": row_cat.raw_code}
            row_total = row_totals.get(row_cat.raw_code, 0)
            for col_cat in col_categories:
                count = count_lookup.get((row_cat.raw_code, col_cat.raw_code), 0)
                if kind == "counts":
                    row[col_cat.label] = count
                elif kind == "rowpct":
                    row[col_cat.label] = (count / row_total * 100.0) if row_total > 0 else 0.0
                else:
                    col_total = colpct_denominators.get(col_cat.raw_code, 0)
                    row[col_cat.label] = (count / col_total * 100.0) if col_total > 0 else 0.0
            if include_grand_total_column:
                if kind == "counts":
                    row["Total"] = row_total
                elif kind == "rowpct":
                    row["Total"] = 100.0 if row_total > 0 else 0.0
                else:
                    row["Total"] = (row_total / base_grand_total * 100.0) if base_grand_total > 0 else 0.0
            rows_out.append(row)

        rows_out.extend(post_rows)

        return {
            "kind": kind,
            "columns": columns,
            "rows": rows_out,
            "header_layout": _build_header_layout(columns),
        }

    tables: List[Dict[str, Any]] = []
    if out_counts:
        tables.append(build("counts"))
    if out_rowpct:
        tables.append(build("rowpct"))
    if out_colpct:
        tables.append(build("colpct"))
    return tables


def _safe_sheet_title(title: str, fallback: str = "Sheet") -> str:
    cleaned = "".join("_" if ch in '[]:*?/\\' else ch for ch in str(title)).strip()
    cleaned = cleaned or fallback
    return cleaned[:31]


def _saved_output_title(result: Dict[str, Any], index: int) -> str:
    row_label = str(result.get("row_label", "Rows")).strip()
    col_label = str(result.get("col_label", "Columns")).strip()
    return f"{index}. {row_label} x {col_label}"


def _renumber_saved_outputs(saved_outputs: List[Dict[str, Any]]) -> None:
    for index, saved in enumerate(saved_outputs, start=1):
        saved["title"] = _saved_output_title(saved["result"], index)


def _move_saved_output_to_index(saved_outputs: List[Dict[str, Any]], current_index: int, new_index: int) -> None:
    item = saved_outputs.pop(current_index)
    saved_outputs.insert(new_index, item)
    _renumber_saved_outputs(saved_outputs)


def _transpose_result(result: Dict[str, Any]) -> Dict[str, Any]:
    transposed_tables: List[Dict[str, Any]] = []
    for table in result.get("tables", []):
        original_columns = list(table.get("columns", []))
        original_rows = list(table.get("rows", []))
        transposed_columns = [str(row.get("__label__", "")) for row in original_rows]
        special_count_columns = [
            str(row.get("__label__", ""))
            for row in original_rows
            if str(row.get("__label__", "")) == "Base"
        ]
        transposed_rows: List[Dict[str, Any]] = []
        for column in original_columns:
            row: Dict[str, Any] = {"__label__": column}
            for original_row in original_rows:
                row[str(original_row.get("__label__", ""))] = original_row.get(column)
            transposed_rows.append(row)
        transposed_tables.append(
            {
                "kind": table.get("kind"),
                "columns": transposed_columns,
                "rows": transposed_rows,
                "header_layout": _build_header_layout(transposed_columns),
                "special_count_columns": special_count_columns,
            }
        )

    output = dict(result)
    output["row_label"] = str(result.get("col_label", "Columns"))
    output["col_label"] = str(result.get("row_label", "Rows"))
    output["row_label_display"] = str(result.get("col_label_display", result.get("col_label", "Columns")))
    output["col_label_display"] = str(result.get("row_label_display", result.get("row_label", "Rows")))
    output["tables"] = transposed_tables
    output["transposed"] = True
    return output


def _render_result_panel(request: Request, job: Dict[str, Any]) -> HTMLResponse:
    return templates.TemplateResponse(
        "_result_panel.html",
        {
            "request": request,
            "job": job,
            "result": job.get("last_result"),
            "pct_suffix": job.get("last_pct_suffix", False),
        },
    )


def _is_partial_request(request: Request) -> bool:
    hx_request = str(request.headers.get("HX-Request", "")).strip().lower()
    requested_with = str(request.headers.get("X-Requested-With", "")).strip().lower()
    return hx_request == "true" or requested_with == "xmlhttprequest"


def _render_error_panel(message: str) -> HTMLResponse:
    safe = str(message).strip() or "Something went wrong while building the crosstab."
    html = (
        '<div class="card">'
        '<div class="card-h"><h3>Couldn\'t Build Preview</h3></div>'
        '<div class="muted">Check the setup below and try again.</div>'
        f'<div class="pillrow" style="margin-top:12px;"><div class="pill"><b>{safe}</b></div></div>'
        "</div>"
    )
    return HTMLResponse(html)


def _get_job_value_dataframe(job: Dict[str, Any]) -> "pd.DataFrame":
    cached = job.get("value_df")
    if cached is not None:
        return cached
    df = _read_excel_object_dataframe(job["value_path"])
    job["value_df"] = df
    return df


def _preview_axis_label(label: Any, fallback: str) -> str:
    text = _normalize_label_text(label)
    if not text:
        return fallback
    text = re.sub(r"^\d+\s*:\s*", "", text)
    text = re.sub(r"^concept\s*\d+\s*:\s*", "", text, flags=re.IGNORECASE)
    text = re.sub(r"\b(Q\d+)_\d+\b", r"\1", text, flags=re.IGNORECASE)
    if len(text) <= 44:
        return text
    q_match = re.search(r"\bQ\d+\b", text, flags=re.IGNORECASE)
    if q_match:
        suffix = text[q_match.start():].strip()
        return suffix if len(suffix) <= 44 else f"{suffix[:41]}..."
    return f"{text[:41]}..."


def _default_banner_column(columns: List[Any], disallowed: Optional[set[str]] = None) -> Optional[str]:
    blocked = disallowed or set()
    for column in columns:
        text = str(column)
        if not text or text == "(none)" or text in blocked:
            continue
        return text
    return None


def _write_result_block(
    ws: Any,
    result: Dict[str, Any],
    pct_suffix: bool,
    start_row: int,
    block_title: Optional[str] = None,
) -> int:
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    fill = PatternFill("solid", fgColor="F2F4F7")

    current_row = start_row
    if block_title:
        ws.cell(row=current_row, column=1, value=block_title).font = Font(bold=True, size=14)
        current_row += 1
        ws.cell(
            row=current_row,
            column=1,
            value=f"Mode: {'Multiple response' if result.get('mode') == 'mr' else 'Single response'} | Column: {result.get('col_label', '')}",
        )
        current_row += 2

    for table in result.get("tables", []):
        kind_to_name = {"counts": "Counts", "rowpct": "Row %", "colpct": "Column %"}
        ws.cell(row=current_row, column=1, value=kind_to_name.get(str(table.get("kind", "")), "Table")).font = header_font
        current_row += 1

        cols = list(table.get("columns", []))
        rows = list(table.get("rows", []))
        special_count_columns = set(table.get("special_count_columns", []))
        header_layout = table.get("header_layout") or _build_header_layout(cols)
        header_row = current_row

        ws.cell(row=header_row, column=1, value=str(result.get("row_label", ""))).font = header_font
        ws.cell(row=header_row, column=1).alignment = center
        ws.cell(row=header_row, column=1).fill = fill

        if not header_layout.get("two_layer"):
            for j, c in enumerate(header_layout.get("bottom_labels", []), start=2):
                cell = ws.cell(row=header_row, column=j, value=c)
                cell.font = header_font
                cell.alignment = center
                cell.fill = PatternFill("solid", fgColor="F8FAFC") if c in special_count_columns else fill
            data_start = header_row + 1
        else:
            running_col = 2
            for group in header_layout.get("top_groups", []):
                cell = ws.cell(row=header_row, column=running_col, value=group["label"])
                cell.font = header_font
                cell.alignment = center
                cell.fill = fill
                span = int(group["span"])
                if span > 1:
                    ws.merge_cells(
                        start_row=header_row,
                        start_column=running_col,
                        end_row=header_row,
                        end_column=running_col + span - 1,
                    )
                running_col += span
            ws.cell(row=header_row + 1, column=1, value="").fill = fill
            for j, label in enumerate(header_layout.get("bottom_labels", []), start=2):
                cell = ws.cell(row=header_row + 1, column=j, value=label)
                cell.font = header_font
                cell.alignment = center
                cell.fill = PatternFill("solid", fgColor="F8FAFC") if label in special_count_columns else fill
            data_start = header_row + 2

        tbl_kind = str(table.get("kind", ""))
        for i, row_data in enumerate(rows, start=data_start):
            row_label = str(row_data.get("__label__", ""))
            is_box_row = bool(row_data.get("__is_box__"))
            ws.cell(row=i, column=1, value=row_label).alignment = left
            for j, c in enumerate(cols, start=2):
                value = row_data.get(c, "")
                if isinstance(value, (int, float)):
                    if tbl_kind in ("rowpct", "colpct") and row_label != "Base" and c not in special_count_columns:
                        vv = int(round(float(value)))
                        value = f"{vv}%" if pct_suffix else vv
                    else:
                        value = int(round(float(value)))
                cell = ws.cell(row=i, column=j, value=value)
                if row_label == "Base" or c in special_count_columns:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="F8FAFC")
                elif is_box_row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="EEF1F5")
            if row_label == "Base":
                ws.cell(row=i, column=1).font = Font(bold=True)
                ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor="F8FAFC")
            elif is_box_row:
                ws.cell(row=i, column=1).font = Font(bold=True)
                ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor="EEF1F5")

        ws.column_dimensions["A"].width = 60
        for j in range(2, len(cols) + 2):
            width = ws.column_dimensions[get_column_letter(j)].width
            ws.column_dimensions[get_column_letter(j)].width = max(width or 0, 14)

        current_row = data_start + len(rows) + 2

    return current_row


def _composite_col(df: pl.DataFrame, col1: str, col2: Optional[str]) -> Tuple[pl.DataFrame, str]:
    if not col2 or col2 == "(none)":
        if col1 not in df.columns:
            raise RuntimeError("Selected column variable not found in dataset.")
        return df, col1
    if col1 not in df.columns or col2 not in df.columns:
        raise RuntimeError("Selected column variable not found in dataset.")
    name = f"{col1} | {col2}"
    if name in df.columns:
        return df, name
    df2 = df.with_columns(
        pl.concat_str(
            [
                pl.col(col1).cast(pl.Utf8, strict=False).fill_null(""),
                pl.lit(" | "),
                pl.col(col2).cast(pl.Utf8, strict=False).fill_null(""),
            ]
        ).alias(name)
    )
    return df2, name


def tabulate_sr(
    df: "pd.DataFrame",
    bundle: MappingBundle,
    row_var: str,
    col_var: str,
    col_var2: Optional[str],
    include_totals: bool,
    include_base: bool,
    out_counts: bool,
    out_rowpct: bool,
    out_colpct: bool,
    include_all_categories: bool = True,
    show_row_codes: bool = False,
    row_sort_mode: str = "code_asc",
    include_top2_box: bool = False,
    include_bottom2_box: bool = False,
    custom_group_label: str = "",
    custom_group_codes: Optional[List[str]] = None,
    custom_groups: Optional[List[Dict[str, Any]]] = None,
    hide_grouped_codes: bool = False,
    weight_col: Optional[str] = None,
) -> Dict[str, Any]:
    effective_row_sort_mode = "value_order" if bundle.source_kind == "codebook" else row_sort_mode
    weights, resolved_weight_col = _resolve_weight_values(df, weight_col)
    row_frame, row_categories = _mapped_series(df, bundle, row_var, include_all_categories=include_all_categories)
    row_codes, _ = _mapped_codes_for_question(df, bundle, row_var)
    col_frame, col_categories, col_label = _prepare_column_dimension(
        df,
        bundle,
        col_var,
        col_var2,
        include_all_categories=include_all_categories,
        total_length_override=len(row_frame.index),
    )
    if not col_var or col_var == "(none)":
        col_codes = ["__total__"] * len(df.index)
    elif not col_var2 or col_var2 == "(none)":
        col_codes, _ = _mapped_codes_for_question(df, bundle, col_var)
    else:
        col_codes = _composite_codes_for_questions(df, bundle, col_var, col_var2)
    rows: List[Dict[str, Any]] = []
    for row_code, col_code, weight in zip(row_codes, col_codes, weights):
        if row_code is None or col_code is None or weight == 0:
            continue
        rows.append({"row_code": row_code, "col_code": col_code, "count": weight})
    counts = pd.DataFrame(rows, columns=["row_code", "col_code", "count"])  # type: ignore[union-attr]
    if not counts.empty:
        counts = counts.groupby(["row_code", "col_code"], dropna=False)["count"].sum().reset_index()
    tables = _format_ordered_tables(
        counts,
        row_categories,
        col_categories,
        include_totals=include_totals,
        include_base=include_base,
        out_counts=out_counts,
        out_rowpct=out_rowpct,
        out_colpct=out_colpct,
        show_row_codes=show_row_codes,
        row_sort_mode=effective_row_sort_mode,
        include_top2_box=include_top2_box,
        include_bottom2_box=include_bottom2_box,
        custom_group_label=custom_group_label,
        custom_group_codes=custom_group_codes,
        custom_groups=custom_groups,
        hide_grouped_codes=hide_grouped_codes,
    )
    return {
        "mode": "sr",
        "row_label": row_var,
        "row_label_display": _preview_axis_label(row_var, "Responses"),
        "col_label": col_label,
        "col_label_display": _preview_axis_label(col_label, "Columns"),
        "tables": tables,
        "validation": [],
        "weight_col": resolved_weight_col,
    }


def tabulate_mr(
    df: "pd.DataFrame",
    bundle: MappingBundle,
    mr_cols: List[str],
    col_var: str,
    col_var2: Optional[str],
    include_totals: bool,
    include_base: bool,
    out_counts: bool,
    out_rowpct: bool,
    out_colpct: bool,
    include_all_categories: bool = True,
    show_row_codes: bool = False,
    row_sort_mode: str = "code_asc",
    mr_detection: Optional[Dict[str, Any]] = None,
    custom_groups: Optional[List[Dict[str, Any]]] = None,
    hide_grouped_codes: bool = False,
    weight_col: Optional[str] = None,
) -> Dict[str, Any]:
    weights, resolved_weight_col = _resolve_weight_values(df, weight_col)
    col_frame, col_categories, col_label = _prepare_column_dimension(
        df,
        bundle,
        col_var,
        col_var2,
        include_all_categories=include_all_categories,
        total_length_override=len(df.index),
    )
    if not col_var or col_var == "(none)":
        col_codes = ["__total__"] * len(df.index)
    elif not col_var2 or col_var2 == "(none)":
        col_codes, _ = _mapped_codes_for_question(df, bundle, col_var)
    else:
        col_codes = _composite_codes_for_questions(df, bundle, col_var, col_var2)

    mr_row_label, row_categories, column_to_category = _build_mr_row_categories(bundle, mr_cols, mr_detection)
    allowed = {"1", "true", "yes", "y", "t"}
    valid_dichotomy = {"0", "1", "true", "false", "yes", "no", "y", "n", "t", "f"}
    rows: List[Dict[str, Any]] = []
    selected_by_question: Dict[str, List[Optional[str]]] = {}
    for question in mr_cols:
        selected_codes = [_normalize_response_value(v) for v in df[question].tolist()]
        selected_by_question[question] = selected_codes
        observed = {code.lower() for code in selected_codes if code is not None}
        invalid = sorted(code for code in observed if code not in valid_dichotomy)
        if invalid:
            shown = ", ".join(invalid[:10])
            raise RuntimeError(f"MR column '{question}' contains non-dichotomy values: {shown}")
        row_category = column_to_category[question]
        for selected, col_code, weight in zip(selected_codes, col_codes, weights):
            if selected is None or selected.lower() not in allowed or col_code is None or weight == 0:
                continue
            rows.append({"row_code": row_category.raw_code, "col_code": col_code, "count": weight})

    respondent_base_rows: List[Tuple[str, float]] = []
    respondent_count = min(
        [len(col_codes), len(weights)] + [len(codes) for codes in selected_by_question.values()]
    ) if selected_by_question else 0
    for idx in range(respondent_count):
        answered = any(
            selected_by_question[question][idx] is not None and selected_by_question[question][idx].lower() in allowed
            for question in mr_cols
        )
        if not answered:
            continue
        col_code = col_codes[idx]
        weight = weights[idx]
        if col_code is None or weight == 0:
            continue
        respondent_base_rows.append((str(col_code), weight))
    base_col_totals: Dict[str, float] = {str(cat.raw_code): 0.0 for cat in col_categories}
    for code, weight in respondent_base_rows:
        base_col_totals[code] = base_col_totals.get(code, 0) + weight
    base_grand_total = sum(weight for _, weight in respondent_base_rows)

    counts = pd.DataFrame(rows, columns=["row_code", "col_code", "count"])  # type: ignore[union-attr]
    if not counts.empty:
        counts = counts.groupby(["row_code", "col_code"], dropna=False)["count"].sum().reset_index()
    resolved_custom_groups: List[Dict[str, Any]] = []
    custom_group_count_overrides: Dict[str, Dict[str, float]] = {}
    for group in list(custom_groups or []):
        resolved_codes: List[str] = []
        for raw_code in list(group.get("codes", []) or []):
            token = str(raw_code).strip()
            if not token:
                continue
            category = column_to_category.get(token)
            if category is not None:
                resolved_codes.append(category.raw_code)
            elif any(row_category.raw_code == token for row_category in row_categories):
                resolved_codes.append(token)
        group_label = str(group.get("label", ""))
        resolved_custom_groups.append({"label": group_label, "codes": resolved_codes})
        if group_label and resolved_codes:
            grouped_counts: Dict[str, float] = {str(cat.raw_code): 0.0 for cat in col_categories}
            included_questions = [
                question for question, category in column_to_category.items()
                if category.raw_code in set(resolved_codes)
            ]
            for idx in range(respondent_count):
                matched = any(
                    selected_by_question[question][idx] is not None and selected_by_question[question][idx].lower() in allowed
                    for question in included_questions
                )
                if not matched:
                    continue
                col_code = col_codes[idx]
                weight = weights[idx]
                if col_code is None or weight == 0:
                    continue
                grouped_counts[str(col_code)] = grouped_counts.get(str(col_code), 0) + weight
            custom_group_count_overrides[group_label] = grouped_counts
    tables = _format_ordered_tables(
        counts,
        row_categories,
        col_categories,
        include_totals=include_totals,
        include_base=include_base,
        out_counts=out_counts,
        out_rowpct=out_rowpct,
        out_colpct=out_colpct,
        show_row_codes=False,
        row_sort_mode="value_order" if bundle.source_kind == "codebook" else "code_asc",
        custom_groups=resolved_custom_groups,
        hide_grouped_codes=hide_grouped_codes,
        base_col_totals_override=base_col_totals,
        base_grand_total_override=base_grand_total,
        colpct_denominator_override=base_col_totals,
        custom_group_count_overrides=custom_group_count_overrides,
    )
    return {
        "mode": "mr",
        "row_label": mr_row_label,
        "row_label_display": _preview_axis_label(mr_row_label, "Responses"),
        "col_label": col_label,
        "col_label_display": _preview_axis_label(col_label, "Columns"),
        "tables": tables,
        "validation": [],
        "weight_col": resolved_weight_col,
    }


@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    return templates.TemplateResponse(
        "index.html",
        {"request": request, "title": APP_TITLE, "version": APP_VERSION, "job": None, "result": None},
    )


@app.post("/upload", response_class=HTMLResponse)
async def upload(
    request: Request,
    file: UploadFile = File(...),
    value_file: UploadFile = File(...),
):
    job_id = str(uuid.uuid4())

    # Per-job temp folder (safe for Koyeb ephemeral FS)
    job_dir = os.path.join(UPLOAD_DIR, job_id)
    os.makedirs(job_dir, exist_ok=True)

    if not file.filename:
        raise HTTPException(status_code=400, detail="Please choose the first workbook to upload.")

    raw_file = await file.read()
    if not raw_file:
        raise HTTPException(status_code=400, detail="The first uploaded workbook is empty.")
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="The first uploaded workbook must be an .xlsx file.")
    if not value_file.filename:
        raise HTTPException(status_code=400, detail="Please choose the second workbook to upload.")

    raw_value_file = await value_file.read()
    if not raw_value_file:
        raise HTTPException(status_code=400, detail="The second uploaded workbook is empty.")
    if not value_file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="The second uploaded workbook must be an .xlsx file.")

    save_path_a = os.path.join(job_dir, f"{job_id}_a_{os.path.basename(file.filename)}")
    with open(save_path_a, "wb") as f:
        f.write(raw_file)
    save_path_b = os.path.join(job_dir, f"{job_id}_b_{os.path.basename(value_file.filename)}")
    with open(save_path_b, "wb") as f:
        f.write(raw_value_file)

    try:
        text_path, value_path, mapping_bundle = _resolve_uploaded_workbook_roles(
            save_path_a,
            file.filename,
            save_path_b,
            value_file.filename,
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

    try:
        df = _read_dataframe(text_path)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    dich_cols = _detect_dichotomy_columns(df)
    mr_detection = detect_mr_groups(df)
    concept_detection = detect_concept_groups([str(column) for column in df.columns])

    # Build column_codes dictionary for filter functionality
    column_codes = {}
    if pd is not None:
        for col in df.columns:
            try:
                unique_values = df[col].dropna().unique()
                column_codes[col] = [str(val) for val in unique_values if val is not None]
            except Exception:
                column_codes[col] = []
    else:
        # Fallback for polars
        for col in df.columns:
            try:
                unique_values = df[col].drop_nulls().unique()
                column_codes[col] = [str(val) for val in unique_values]
            except Exception:
                column_codes[col] = []

    columns = [str(column) for column in df.columns]
    filter_value_options = _build_filter_value_options(df, mapping_bundle)

    JOBS[job_id] = {
        "id": job_id,
        "name": os.path.basename(text_path),
        "path": text_path,
        "value_path": value_path,
        "value_df": None,
        "mapping_bundle": mapping_bundle,
        "groupable_question_options": _groupable_question_options(mapping_bundle),
        "columns": columns,
        "column_display_labels": _column_display_labels(columns),
        "column_codes": column_codes,
        "filter_value_options": filter_value_options,
        "dich_cols": dich_cols,
        "mr_detection": mr_detection,
        "concept_detection": concept_detection,
        "saved_outputs": [],
        "last_saved_message": None,
        "saved_modal_open": False,
    }

    if _is_partial_request(request):
        return templates.TemplateResponse(
            "_dataset_panel.html",
            {"request": request, "title": APP_TITLE, "job": JOBS[job_id]},
        )
    return templates.TemplateResponse(
        "index.html",
        {"request": request, "title": APP_TITLE, "version": APP_VERSION, "job": JOBS[job_id], "result": None},
    )

@app.post("/run", response_class=HTMLResponse)
def run(
    request: Request,
    job_id: str = Form(...),
    row_var: str = Form("(none)"),
    col_var: str = Form("(none)"),
    col_var2: str = Form("(none)"),
    qtype: str = Form("sr"),
    mr_cols: List[str] = Form([]),
    weight_col: str = Form("(none)"),
    out_counts: str | None = Form(None),
    out_rowpct: str | None = Form(None),
    out_colpct: str | None = Form(None),
    transpose_table: str | None = Form(None),
    pct_suffix: str | None = Form(None),
    show_row_codes: str | None = Form(None),
    compare_concepts: str | None = Form(None),
    include_top2_box: str | None = Form(None),
    include_bottom2_box: str | None = Form(None),
    hide_grouped_codes: str | None = Form(None),
    custom_group_definitions: str = Form(""),
    custom_group_label: str = Form(""),
    custom_group_codes: List[str] = Form([]),
    filter_data: str = Form(""),
    row_sort_mode: str = Form("code_asc"),
    include_base: str | None = Form(None),
    include_totals: str | None = Form(None),
):
    if job_id not in JOBS:
        return _render_error_panel("Invalid job id.")

    job = JOBS[job_id]
    mapping_bundle = job.get("mapping_bundle")
    if mapping_bundle is None:
        return _render_error_panel("This crosstab requires an uploaded text workbook paired with Value.xlsx.")
    if pd is None:
        return _render_error_panel("Pandas is required for ordered tabulation.")
    df = _get_job_value_dataframe(job)
    parsed_filter_data: Dict[str, Any] = {}
    filter_preview = {"human_readable": "No filters applied", "structured": "{}"}
    filter_warnings: List[str] = []
    if str(filter_data).strip():
        try:
            parsed_filter_data = json.loads(str(filter_data))
        except Exception:
            return _render_error_panel("Invalid filter configuration.")
        if parsed_filter_data.get("active"):
            filter_warnings = _filter_logic_warnings(parsed_filter_data)
            try:
                df = apply_filters_to_dataframe(df, parsed_filter_data)
            except Exception as e:
                return _render_error_panel(f"Filter error: {e}")
            filter_preview = generate_filter_preview(parsed_filter_data)

    # Populate column_codes if missing (for existing jobs)
    if "column_codes" not in job or not job["column_codes"]:
        column_codes = {}
        if pd is not None:
            for col in df.columns:
                try:
                    unique_values = df[col].dropna().unique()
                    column_codes[col] = [str(val) for val in unique_values if val is not None]
                except Exception:
                    column_codes[col] = []
        else:
            # Fallback for polars
            for col in df.columns:
                try:
                    unique_values = df[col].drop_nulls().unique()
                    column_codes[col] = [str(val) for val in unique_values]
                except Exception:
                    column_codes[col] = []
        job["column_codes"] = column_codes
    if "filter_value_options" not in job or not job["filter_value_options"]:
        job["filter_value_options"] = _build_filter_value_options(df, mapping_bundle)

    f_counts = out_counts is not None
    f_rowpct = out_rowpct is not None
    f_colpct = out_colpct is not None
    f_transpose = transpose_table is not None
    f_base = include_base is not None
    f_totals = include_totals is not None
    f_show_row_codes = show_row_codes is not None
    f_compare_concepts = compare_concepts is not None
    f_top2_box = include_top2_box is not None
    f_bottom2_box = include_bottom2_box is not None
    f_hide_grouped_codes = hide_grouped_codes is not None
    parsed_custom_groups = _parse_custom_group_definitions(
        custom_group_definitions,
        single_label=custom_group_label,
        single_codes=custom_group_codes,
    )
    if row_sort_mode not in {"code_asc", "code_desc"}:
        return _render_error_panel("Invalid row sort mode.")

    if not any([f_counts, f_rowpct, f_colpct]):
        return _render_error_panel("Select at least one output option.")

    try:
        if qtype == "mr":
            if not mr_cols:
                return _render_error_panel("No MR columns selected.")
            result = tabulate_mr(
                df,
                bundle=mapping_bundle,
                mr_cols=mr_cols,
                col_var=col_var,
                col_var2=None if col_var in {"", "(none)"} or col_var2 == "(none)" else col_var2,
                include_totals=f_totals, include_base=f_base,
                out_counts=f_counts, out_rowpct=f_rowpct, out_colpct=f_colpct,
                show_row_codes=f_show_row_codes,
                row_sort_mode="code_asc",
                mr_detection=job.get("mr_detection"),
                custom_groups=parsed_custom_groups,
                hide_grouped_codes=f_hide_grouped_codes,
                weight_col=weight_col,
            )
        else:
            if row_var in {"", "(none)"}:
                return _render_error_panel("Select a row variable.")
            if f_compare_concepts:
                concept_detection = job.get("concept_detection") or {}
                concept_group = (concept_detection.get("member_to_group") or {}).get(row_var)
                if concept_group is None:
                    return _render_error_panel(
                        "Concept comparison is only available when the selected row variable belongs to a detected concept family."
                    )
                result = tabulate_sr_concept_comparison(
                    df,
                    bundle=mapping_bundle,
                    concept_group=concept_group,
                    top_banner_var=None if col_var == "(none)" else col_var,
                    include_totals=f_totals,
                    include_base=f_base,
                    out_counts=f_counts,
                    out_rowpct=f_rowpct,
                    out_colpct=f_colpct,
                    show_row_codes=f_show_row_codes,
                    row_sort_mode=row_sort_mode,
                    include_top2_box=f_top2_box,
                    include_bottom2_box=f_bottom2_box,
                    custom_group_label=custom_group_label,
                    custom_group_codes=custom_group_codes,
                    custom_groups=parsed_custom_groups,
                    hide_grouped_codes=f_hide_grouped_codes,
                    weight_col=weight_col,
                )
            else:
                result = tabulate_sr(
                    df,
                    bundle=mapping_bundle,
                    row_var=row_var,
                    col_var=col_var,
                    col_var2=None if col_var in {"", "(none)"} or col_var2 == "(none)" else col_var2,
                    include_totals=f_totals, include_base=f_base,
                    out_counts=f_counts, out_rowpct=f_rowpct, out_colpct=f_colpct,
                    show_row_codes=f_show_row_codes,
                    row_sort_mode=row_sort_mode,
                    include_top2_box=f_top2_box,
                    include_bottom2_box=f_bottom2_box,
                    custom_group_label=custom_group_label,
                    custom_group_codes=custom_group_codes,
                    custom_groups=parsed_custom_groups,
                    hide_grouped_codes=f_hide_grouped_codes,
                    weight_col=weight_col,
                )
                result["row_var"] = row_var
        result["col_var"] = col_var
        result["col_var2"] = col_var2
        result["mode"] = qtype
        result["weight_col"] = weight_col
        result["show_row_codes"] = f_show_row_codes
        result["include_top2_box"] = f_top2_box
        result["include_bottom2_box"] = f_bottom2_box
        result["custom_group_definitions"] = custom_group_definitions
        result["custom_group_label"] = custom_group_label
        result["row_sort_mode"] = row_sort_mode
        result["hide_grouped_codes"] = f_hide_grouped_codes
        result["mr_cols"] = mr_cols if qtype == "mr" else []
        result["filter_data"] = parsed_filter_data
        result["filter_preview"] = filter_preview
        result["filter_active"] = bool(parsed_filter_data.get("active"))
        if filter_warnings:
            result["validation"] = list(result.get("validation") or []) + filter_warnings
        if f_transpose:
            result = _transpose_result(result)
        result["mapping_assumptions"] = list(mapping_bundle.assumptions)
        result["mapping_value_path"] = mapping_bundle.value_path
    except Exception as e:
        return _render_error_panel(str(e))

    job['last_result'] = result
    job['last_pct_suffix'] = (pct_suffix is not None)
    job['last_saved_message'] = None
    job['saved_modal_open'] = False
    return _render_result_panel(request, job)


@app.post("/save-output/{job_id}", response_class=HTMLResponse)
def save_output(request: Request, job_id: str):
    job = JOBS.get(job_id)
    if not job or "last_result" not in job:
        return HTMLResponse("No output available to save.", status_code=400)

    saved_outputs = job.setdefault("saved_outputs", [])
    snapshot = {
        "title": _saved_output_title(job["last_result"], len(saved_outputs) + 1),
        "result": deepcopy(job["last_result"]),
        "pct_suffix": bool(job.get("last_pct_suffix", False)),
    }
    saved_outputs.append(snapshot)
    job["last_saved_message"] = f"Saved output #{len(saved_outputs)} for later group export."
    job["saved_modal_open"] = False
    return _render_result_panel(request, job)


@app.post("/delete-saved-output/{job_id}/{saved_index}", response_class=HTMLResponse)
def delete_saved_output(request: Request, job_id: str, saved_index: int):
    job = JOBS.get(job_id)
    if not job:
        return HTMLResponse("Invalid job id", status_code=400)

    saved_outputs = job.get("saved_outputs") or []
    if saved_index < 0 or saved_index >= len(saved_outputs):
        return HTMLResponse("Saved output not found.", status_code=400)

    removed = saved_outputs.pop(saved_index)
    _renumber_saved_outputs(saved_outputs)
    job["last_saved_message"] = f"Deleted saved output: {removed['title']}"
    job["saved_modal_open"] = True
    return _render_result_panel(request, job)


@app.post("/move-saved-output/{job_id}/{saved_index}", response_class=HTMLResponse)
def move_saved_output(request: Request, job_id: str, saved_index: int, direction: str = Form(...)):
    job = JOBS.get(job_id)
    if not job:
        return HTMLResponse("Invalid job id", status_code=400)

    saved_outputs = job.get("saved_outputs") or []
    if saved_index < 0 or saved_index >= len(saved_outputs):
        return HTMLResponse("Saved output not found.", status_code=400)
    if direction not in {"up", "down"}:
        return HTMLResponse("Invalid move direction.", status_code=400)

    new_index = saved_index - 1 if direction == "up" else saved_index + 1
    if new_index < 0 or new_index >= len(saved_outputs):
        return _render_result_panel(request, job)

    _move_saved_output_to_index(saved_outputs, saved_index, new_index)
    job["last_saved_message"] = f"Moved saved output to position {new_index + 1}."
    job["saved_modal_open"] = True
    return _render_result_panel(request, job)


@app.post("/renumber-saved-output/{job_id}/{saved_index}", response_class=HTMLResponse)
def renumber_saved_output(request: Request, job_id: str, saved_index: int, target_position: int = Form(...)):
    job = JOBS.get(job_id)
    if not job:
        return HTMLResponse("Invalid job id", status_code=400)

    saved_outputs = job.get("saved_outputs") or []
    if saved_index < 0 or saved_index >= len(saved_outputs):
        return HTMLResponse("Saved output not found.", status_code=400)
    if not saved_outputs:
        return HTMLResponse("No saved outputs available.", status_code=400)

    bounded_position = max(1, min(int(target_position), len(saved_outputs)))
    new_index = bounded_position - 1
    if new_index != saved_index:
        _move_saved_output_to_index(saved_outputs, saved_index, new_index)
        job["last_saved_message"] = f"Moved saved output to position {bounded_position}."
    else:
        job["last_saved_message"] = f"Saved output is already at position {bounded_position}."
    job["saved_modal_open"] = True
    return _render_result_panel(request, job)


@app.post("/resequence-saved-outputs/{job_id}", response_class=HTMLResponse)
async def resequence_saved_outputs(request: Request, job_id: str):
    job = JOBS.get(job_id)
    if not job:
        return HTMLResponse("Invalid job id", status_code=400)

    saved_outputs = job.get("saved_outputs") or []
    if not saved_outputs:
        return HTMLResponse("No saved outputs available.", status_code=400)

    form = await request.form()
    indexed_items: List[Tuple[int, int, Dict[str, Any]]] = []
    for index, saved in enumerate(saved_outputs):
        raw_target = form.get(f"seq_{index}", index + 1)
        try:
            target_position = int(str(raw_target))
        except ValueError:
            target_position = index + 1
        bounded_position = max(1, min(target_position, len(saved_outputs)))
        indexed_items.append((bounded_position, index, saved))

    indexed_items.sort(key=lambda item: (item[0], item[1]))
    reordered = [saved for _, _, saved in indexed_items]
    saved_outputs[:] = reordered
    _renumber_saved_outputs(saved_outputs)
    job["last_saved_message"] = "Applied saved output sequence changes."
    job["saved_modal_open"] = True
    return _render_result_panel(request, job)




@app.get("/export-saved/{job_id}")
def export_saved(job_id: str, mode: str = "single"):
    if job_id not in JOBS:
        return HTMLResponse("Job not found.", status_code=404)

    job = JOBS[job_id]
    saved_outputs = job.get("saved_outputs", [])
    if not saved_outputs:
        return HTMLResponse("No result to export for this job.", status_code=400)
    if mode not in {"individual", "single"}:
        return HTMLResponse("Invalid group export mode.", status_code=400)

    wb = Workbook()
    ws0 = wb.active
    wb.remove(ws0)

    if mode == "individual":
        for index, saved in enumerate(saved_outputs, start=1):
            ws = wb.create_sheet(title=_safe_sheet_title(saved["title"], fallback=f"Output {index}"))
            _write_result_block(
                ws,
                saved["result"],
                pct_suffix=bool(saved.get("pct_suffix", False)),
                start_row=1,
                block_title=saved["title"],
            )
            ws.freeze_panes = ws["B4"]
    else:
        ws = wb.create_sheet(title="Group Export")
        current_row = 1
        for saved in saved_outputs:
            current_row = _write_result_block(
                ws,
                saved["result"],
                pct_suffix=bool(saved.get("pct_suffix", False)),
                start_row=current_row,
                block_title=saved["title"],
            )
        ws.freeze_panes = ws["B4"]

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"group_export_{job_id}_{mode}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.get("/export/{job_id}")
def export_excel(job_id: str):
    job = JOBS.get(job_id)
    if not job or "last_result" not in job:
        return HTMLResponse("No result to export for this job.", status_code=400)

    result = job["last_result"]
    # Mirror the preview behaviour: if the user chose to show the percent sign,
    # append "%" on Row% / Column% exports (but not on base rows).
    pct_suffix = bool(job.get("last_pct_suffix", False))
    weighted_counts = bool(result.get("weight_col"))
    wb = Workbook()
    # remove default sheet
    ws0 = wb.active
    wb.remove(ws0)

    def write_table(sheet_name: str, table: dict):
        ws = wb.create_sheet(title=sheet_name[:31])
        cols = list(table["columns"])
        rows = list(table["rows"])

        # header styling
        header_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="top", wrap_text=True)
        fill = PatternFill("solid", fgColor="F2F4F7")

        header_layout = table.get("header_layout") or _build_header_layout(cols)
        start_row = 1
        # Row label column header
        ws.cell(row=start_row, column=1, value=str(result.get("row_label",""))).font = header_font
        ws.cell(row=start_row, column=1).alignment = center
        ws.cell(row=start_row, column=1).fill = fill

        if not header_layout.get("two_layer"):
            # single header row
            for j,c in enumerate(header_layout.get("bottom_labels", []), start=2):
                cell = ws.cell(row=start_row, column=j, value=c)
                cell.font = header_font
                cell.alignment = center
                cell.fill = fill
            data_start = start_row + 1
        else:
            # two header rows
            # top row merges contiguous equal labels
            running_col = 2
            for group in header_layout.get("top_groups", []):
                cell = ws.cell(row=start_row, column=running_col, value=group["label"])
                cell.font = header_font
                cell.alignment = center
                cell.fill = fill
                span = int(group["span"])
                if span > 1:
                    ws.merge_cells(start_row=start_row, start_column=running_col, end_row=start_row, end_column=running_col + span - 1)
                running_col += span
            # second header row
            ws.cell(row=start_row+1, column=1, value="").fill = fill
            for j,label in enumerate(header_layout.get("bottom_labels", []), start=2):
                cell = ws.cell(row=start_row+1, column=j, value=label)
                cell.font = header_font
                cell.alignment = center
                cell.fill = fill
            data_start = start_row + 2

        # rows
        tbl_kind = str(table.get("kind", ""))
        for i,r in enumerate(rows, start=data_start):
            row_label = str(r.get("__label__",""))
            is_box_row = bool(r.get("__is_box__"))
            ws.cell(row=i, column=1, value=row_label).alignment = left
            for j,c in enumerate(cols, start=2):
                v = r.get(c, "")
                # Export formatting: whole numbers only.
                # - Counts/Base: integers unless weighted
                # - Row%/Col%: integer percentages
                if isinstance(v, (int, float)):
                    if tbl_kind in ("rowpct", "colpct") and row_label != "Base":
                        vv = int(round(float(v)))
                        v = f"{vv}%" if pct_suffix else vv
                    else:
                        v = round(float(v), 2) if weighted_counts else int(round(float(v)))
                cell = ws.cell(row=i, column=j, value=v)
                # base row highlight
                if row_label == "Base":
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="F8FAFC")
                elif is_box_row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="EEF1F5")
            if row_label == "Base":
                ws.cell(row=i, column=1).font = Font(bold=True)
                ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor="F8FAFC")
            elif is_box_row:
                ws.cell(row=i, column=1).font = Font(bold=True)
                ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor="EEF1F5")

        # column widths
        ws.column_dimensions["A"].width = 60
        for j in range(2, len(cols)+2):
            ws.column_dimensions[get_column_letter(j)].width = 14

        ws.freeze_panes = ws["B{}".format(data_start)]
        return ws

    kind_to_name = {"counts":"Counts", "rowpct":"Row %", "colpct":"Column %"}
    for t in result.get("tables", []):
        sheet = kind_to_name.get(t.get("kind",""), str(t.get("kind","Table")))
        write_table(sheet, t)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"crosstab_{job_id}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


# Filter System Functions

@dataclass
class FilterCondition:
    variable: str
    operator: str
    value: str
    variable_type: str = "text"


@dataclass
class FilterGroup:
    operator: str  # "AND" or "OR"
    conditions: List[FilterCondition]
    nested_groups: List['FilterGroup'] = None


def _detect_variable_type(df: "pd.DataFrame", column: str) -> str:
    """Detect the type of a variable for filtering purposes."""
    if column not in df.columns:
        return "text"
    
    series = df[column]
    
    # Check if it's numeric
    if pd.api.types.is_numeric_dtype(series):
        return "numeric"
    
    # Check if it's boolean-like
    if _is_dichotomy_series(series):
        return "boolean"
    
    # Check if it has limited unique values (categorical)
    unique_count = series.dropna().nunique()
    if unique_count <= 20 and unique_count < len(series) * 0.5:
        return "categorical"
    
    return "text"


def _get_operators_for_type(var_type: str) -> List[str]:
    """Get available operators for a variable type."""
    if var_type == "numeric":
        return ["=", "!=", ">", ">=", "<=", "between"]
    elif var_type == "boolean":
        return ["=", "!="]
    elif var_type == "categorical":
        return ["=", "!=", "is", "is not", "in list"]
    else:  # text
        return ["=", "!=", "contains", "starts with", "ends with"]


def _apply_filter_condition(df: "pd.DataFrame", condition: FilterCondition) -> "pd.DataFrame":
    """Apply a single filter condition to a dataframe."""
    if condition.variable not in df.columns:
        return df
    
    series = df[condition.variable]
    normalized_value = str(condition.value).strip()

    def _normalized_text_mask(expected: str, negate: bool = False):
        series_text = series.astype(str).str.strip()
        mask = series_text == expected
        return ~mask if negate else mask

    def _normalized_numeric_mask(expected: str, negate: bool = False):
        numeric_series = pd.to_numeric(series, errors='coerce')
        expected_number = pd.to_numeric(pd.Series([expected]), errors='coerce').iloc[0]
        if pd.isna(expected_number):
            return _normalized_text_mask(expected, negate=negate)
        mask = numeric_series == expected_number
        return ~mask if negate else mask
    
    if condition.operator == "=":
        if pd.api.types.is_numeric_dtype(series):
            mask = _normalized_numeric_mask(normalized_value)
        else:
            mask = _normalized_text_mask(normalized_value)
    elif condition.operator == "!=":
        if pd.api.types.is_numeric_dtype(series):
            mask = _normalized_numeric_mask(normalized_value, negate=True)
        else:
            mask = _normalized_text_mask(normalized_value, negate=True)
    elif condition.operator == ">":
        mask = pd.to_numeric(series, errors='coerce') > float(normalized_value)
    elif condition.operator == ">=":
        mask = pd.to_numeric(series, errors='coerce') >= float(normalized_value)
    elif condition.operator == "<":
        mask = pd.to_numeric(series, errors='coerce') < float(normalized_value)
    elif condition.operator == "<=":
        mask = pd.to_numeric(series, errors='coerce') <= float(normalized_value)
    elif condition.operator == "between":
        try:
            min_val, max_val = normalized_value.split(",")
            mask = pd.to_numeric(series, errors='coerce').between(float(min_val), float(max_val))
        except:
            mask = pd.Series([False] * len(series))
    elif condition.operator == "contains":
        mask = series.astype(str).str.contains(normalized_value, na=False)
    elif condition.operator == "starts with":
        mask = series.astype(str).str.startswith(normalized_value, na=False)
    elif condition.operator == "ends with":
        mask = series.astype(str).str.endswith(normalized_value, na=False)
    elif condition.operator == "is":
        if pd.api.types.is_numeric_dtype(series):
            mask = _normalized_numeric_mask(normalized_value)
        else:
            mask = _normalized_text_mask(normalized_value)
    elif condition.operator == "is not":
        if pd.api.types.is_numeric_dtype(series):
            mask = _normalized_numeric_mask(normalized_value, negate=True)
        else:
            mask = _normalized_text_mask(normalized_value, negate=True)
    elif condition.operator == "in list":
        values = [v.strip() for v in normalized_value.split(",") if v.strip()]
        if pd.api.types.is_numeric_dtype(series):
            numeric_values = pd.to_numeric(pd.Series(values), errors='coerce').dropna().tolist()
            if numeric_values:
                mask = pd.to_numeric(series, errors='coerce').isin(numeric_values)
            else:
                mask = series.astype(str).str.strip().isin(values)
        else:
            mask = series.astype(str).str.strip().isin(values)
    else:
        mask = pd.Series([True] * len(series))
    
    return df[mask]


def _apply_filter_group(df: "pd.DataFrame", group: FilterGroup) -> "pd.DataFrame":
    """Apply a filter group (with conditions and nested groups) to a dataframe."""
    if not group.conditions and not group.nested_groups:
        return df
    
    results = []
    
    # Apply conditions in this group
    if group.conditions:
        if group.operator == "AND":
            condition_result = df.copy()
            for condition in group.conditions:
                condition_result = _apply_filter_condition(condition_result, condition)
            results.append(condition_result)
        else:
            for condition in group.conditions:
                results.append(_apply_filter_condition(df, condition))
    
    # Apply nested groups
    if group.nested_groups:
        for nested_group in group.nested_groups:
            nested_result = _apply_filter_group(df, nested_group)
            results.append(nested_result)
    
    if not results:
        return df
    
    if group.operator == "AND":
        # For AND, intersect all results
        result = results[0]
        for other_result in results[1:]:
            result = pd.merge(result, other_result, how='inner', on=list(df.columns))
        return result
    else:  # OR
        # For OR, union matching respondent rows without collapsing duplicate values.
        combined = pd.concat(results, axis=0)
        combined = combined.loc[~combined.index.duplicated(keep='first')]
        return combined.sort_index()


def parse_quick_filter(conditions_data: List[Dict[str, str]], match_type: str) -> FilterGroup:
    """Parse quick filter data into a FilterGroup."""
    conditions = []
    for cond_data in conditions_data:
        if not cond_data.get('variable') or not cond_data.get('operator'):
            continue
        conditions.append(FilterCondition(
            variable=cond_data['variable'],
            operator=cond_data['operator'],
            value=cond_data.get('value', ''),
            variable_type=cond_data.get('variable_type', 'text')
        ))

    group_operator = "OR" if str(match_type).lower() == "any" else "AND"
    return FilterGroup(operator=group_operator, conditions=conditions)


def parse_advanced_filter(groups_data: List[Dict[str, Any]]) -> FilterGroup:
    """Parse advanced filter data into nested FilterGroups."""
    groups = []
    for group_data in groups_data:
        conditions = []
        for cond_data in group_data.get('conditions', []):
            if not cond_data.get('variable') or not cond_data.get('operator'):
                continue
            conditions.append(FilterCondition(
                variable=cond_data['variable'],
                operator=cond_data['operator'],
                value=cond_data.get('value', ''),
                variable_type=cond_data.get('variable_type', 'text')
            ))
        
        nested_groups = []
        for nested_data in group_data.get('nested_groups', []):
            nested_groups.append(parse_advanced_filter([nested_data]))
        
        groups.append(FilterGroup(
            operator=group_data.get('operator') or group_data.get('group_operator', 'AND'),
            conditions=conditions,
            nested_groups=nested_groups
        ))
    
    # If multiple groups, combine them with AND
    if len(groups) == 1:
        return groups[0]
    else:
        return FilterGroup(operator="AND", conditions=[], nested_groups=groups)


def _filter_logic_warnings(filter_data: Dict[str, Any]) -> List[str]:
    warnings: List[str] = []
    if not filter_data or not filter_data.get("active"):
        return warnings

    def check_conditions(conditions: List[Dict[str, Any]], join_operator: str) -> None:
        if str(join_operator).upper() != "AND":
            return
        equals_by_variable: Dict[str, set[str]] = {}
        for cond in conditions:
            variable = str(cond.get("variable", "") or "")
            operator = str(cond.get("operator", "") or "").strip().lower()
            value = str(cond.get("value", "") or "").strip()
            if not variable or not value:
                continue
            if operator not in {"=", "is"}:
                continue
            equals_by_variable.setdefault(variable, set()).add(value)
        for variable, values in equals_by_variable.items():
            if len(values) > 1:
                rendered = ", ".join(sorted(values))
                warnings.append(
                    f"Filter warning: {variable} is checked against multiple exact values ({rendered}) with AND, so a single-response variable will return 0 rows."
                )

    if filter_data.get("type", "quick") == "quick":
        match_type = "AND" if str(filter_data.get("match_type", "all")).lower() == "all" else "OR"
        check_conditions(list(filter_data.get("conditions", []) or []), match_type)
    else:
        for group in list(filter_data.get("groups", []) or []):
            operator = str(group.get("operator") or group.get("group_operator") or "AND")
            check_conditions(list(group.get("conditions", []) or []), operator)
    return warnings


def apply_filters_to_dataframe(df: "pd.DataFrame", filter_data: Dict[str, Any]) -> "pd.DataFrame":
    """Apply filters to a dataframe based on filter data."""
    if not filter_data or not filter_data.get('active'):
        return df
    
    filter_type = filter_data.get('type', 'quick')
    
    if filter_type == 'quick':
        conditions = filter_data.get('conditions', [])
        match_type = filter_data.get('match_type', 'all')
        if not conditions:
            return df
        
        filter_group = parse_quick_filter(conditions, match_type)
        return _apply_filter_group(df, filter_group)
    
    elif filter_type == 'advanced':
        groups = filter_data.get('groups', [])
        if not groups:
            return df
        
        filter_group = parse_advanced_filter(groups)
        return _apply_filter_group(df, filter_group)
    
    return df


def generate_filter_preview(filter_data: Dict[str, Any]) -> Dict[str, str]:
    """Generate human-readable and structured previews of filter logic."""
    if not filter_data or not filter_data.get('active'):
        return {
            "human_readable": "No filters applied",
            "structured": "{}"
        }
    
    filter_type = filter_data.get('type', 'quick')
    
    if filter_type == 'quick':
        conditions = filter_data.get('conditions', [])
        match_type = filter_data.get('match_type', 'all')
        
        if not conditions:
            return {
                "human_readable": "No filters defined",
                "structured": "{}"
            }
        
        condition_texts = []
        for cond in conditions:
            var = cond.get('variable', '')
            op = cond.get('operator', '')
            val = cond.get('value', '')
            condition_texts.append(f"{var} {op} {val}")
        
        if match_type == 'all':
            human_readable = " AND ".join(condition_texts)
        else:
            human_readable = " OR ".join(condition_texts)
        
        structured = {
            "type": "quick",
            "match_type": match_type,
            "conditions": conditions
        }
        
    else:  # advanced
        groups = filter_data.get('groups', [])
        
        if not groups:
            return {
                "human_readable": "No filters defined",
                "structured": "{}"
            }
        
        group_texts = []
        structured_groups = []
        
        for group in groups:
            operator = group.get('operator') or group.get('group_operator', 'AND')
            conditions = group.get('conditions', [])
            nested_groups = group.get('nested_groups', [])
            
            condition_texts = []
            structured_conditions = []
            
            for cond in conditions:
                var = cond.get('variable', '')
                op = cond.get('operator', '')
                val = cond.get('value', '')
                condition_texts.append(f"{var} {op} {val}")
                structured_conditions.append({
                    "variable": var,
                    "operator": op,
                    "value": val
                })
            
            group_text = f"({f' {operator} '.join(condition_texts)})" if condition_texts else ""
            group_texts.append(group_text)
            
            structured_groups.append({
                "operator": operator,
                "conditions": structured_conditions,
                "nested_groups": nested_groups
            })
        
        human_readable = " AND ".join([gt for gt in group_texts if gt])
        structured = {
            "type": "advanced",
            "groups": structured_groups
        }
    
    return {
        "human_readable": human_readable,
        "structured": str(structured).replace("'", '"')
    }


@app.post("/get-variable-info")
async def get_variable_info(request: Request, job_id: str = Form(...), variable: str = Form(...)):
    """Get information about a variable including type and available operators."""
    if job_id not in JOBS:
        return JSONResponse({"error": "Invalid job ID"}, status_code=404)
    
    job = JOBS[job_id]
    df = _get_job_value_dataframe(job)
    
    var_type = _detect_variable_type(df, variable)
    operators = _get_operators_for_type(var_type)
    
    # Get unique values for categorical/text variables
    unique_values = []
    value_options = []
    mapping_bundle = job.get("mapping_bundle")
    if mapping_bundle is not None and variable in getattr(mapping_bundle, "questions", {}):
        try:
            mapping = mapping_bundle.questions[variable]
            value_options = [
                {
                    "value": str(category.raw_code),
                    "label": f"{category.raw_code} - {category.label}",
                }
                for category in mapping.categories
            ]
            unique_values = [str(category.raw_code) for category in mapping.categories]
        except Exception:
            value_options = []
    if var_type in ["categorical", "text", "boolean"]:
        try:
            if not unique_values:
                unique_vals = df[variable].dropna().unique()
                unique_values = [str(val) for val in unique_vals[:50]]  # Limit to 50 values
            if not value_options:
                value_options = [{"value": str(val), "label": str(val)} for val in unique_values]
        except Exception:
            pass
    
    return JSONResponse({
        "type": var_type,
        "operators": operators,
        "unique_values": unique_values,
        "value_options": value_options or job.get("filter_value_options", {}).get(variable, []),
    })


@app.post("/preview-filter")
async def preview_filter_logic(request: Request, filter_data: str = Form(...)):
    """Preview filter logic without applying it."""
    try:
        import json
        filter_dict = json.loads(filter_data)
        preview = generate_filter_preview(filter_dict)
        return JSONResponse(preview)
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=400)


@app.get("/test-filter", response_class=HTMLResponse)
async def test_filter_panel(request: Request):
    """Test endpoint to verify filter panel is working."""
    # Use absolute path for template
    template_path = os.path.join(BASE_DIR, "templates", "test_filter.html")
    if not os.path.exists(template_path):
        return HTMLResponse("<h1>Template not found</h1>", status_code=404)
    
    with open(template_path, 'r', encoding='utf-8') as f:
        template_content = f.read()
    
    # Simple template rendering without Jinja2 for this test
    from jinja2 import Template
    template = Template(template_content)
    rendered = template.render({
        "request": request, 
        "title": "Filter Test", 
        "cols": ["test_var1", "test_var2", "test_var3"]
    })
    
    return HTMLResponse(rendered)


@app.get("/variables")
async def get_variables(dataset_id: str):
    """Get variables and metadata for a dataset."""
    if dataset_id not in JOBS:
        return JSONResponse({"error": "Invalid job ID"}, status_code=404)
    
    job = JOBS[dataset_id]
    return JSONResponse({
        "columns": list(job.get("columns", [])),
        "dich_cols": list(job.get("dich_cols", [])),
        "mr_detection": job.get("mr_detection", {}),
        "concept_detection": job.get("concept_detection", {}),
        "column_codes": job.get("column_codes", {})
    })


@app.post("/crosstab", response_class=JSONResponse)
async def crosstab_endpoint(request: Request, payload: Dict[str, Any]):
    """Generate crosstab with optional filtering."""
    try:
        job_id = payload.get("dataset_id")
        if not job_id or job_id not in JOBS:
            return JSONResponse({"error": "Invalid job ID"}, status_code=404)
        
        job = JOBS[job_id]
        mapping_bundle = job.get("mapping_bundle")
        if mapping_bundle is None:
            return JSONResponse({"error": "This crosstab requires an uploaded text workbook paired with Value.xlsx."}, status_code=400)
        
        if pd is None:
            return JSONResponse({"error": "Pandas is required for ordered tabulation."}, status_code=400)
        
        df = _get_job_value_dataframe(job)
        
        # Apply filters if provided
        filter_data = payload.get("filter_data")
        if filter_data and filter_data.get("active"):
            df = apply_filters_to_dataframe(df, filter_data)
        
        # Extract parameters
        row_var = payload.get("row", "(none)")
        col_var = payload.get("col1", "(none)")
        col_var2 = payload.get("col2", "(none)")
        weight_col = payload.get("weight", "(none)")
        qtype = "mr" if payload.get("mr_mode") else "sr"
        mr_cols = []  # TODO: extract from payload if needed
        
        # Parse boolean options
        include_counts = payload.get("include_counts", True)
        include_row_pct = payload.get("include_row_pct", True)
        include_col_pct = payload.get("include_col_pct", True)
        include_totals = payload.get("include_totals", True)
        include_base = payload.get("include_base", True)
        transpose_table = payload.get("transpose_table", False)
        show_row_codes = payload.get("show_row_codes", False)
        compare_concepts = payload.get("compare_concepts", False)
        include_top2_box = payload.get("include_top2_box", False)
        include_bottom2_box = payload.get("include_bottom2_box", False)
        hide_grouped_codes = payload.get("hide_grouped_codes", False)
        row_sort_mode = payload.get("row_sort_mode", "code_asc")
        mr_value = payload.get("mr_value", "1")
        custom_group_definitions = payload.get("custom_group_definitions", "")
        custom_group_label = payload.get("custom_group_label", "")
        custom_group_codes = payload.get("custom_group_codes", [])
        
        # Validate required parameters
        if not any([include_counts, include_row_pct, include_col_pct]):
            return JSONResponse({"error": "Select at least one output option."}, status_code=400)
        
        # Generate crosstab
        if qtype == "mr":
            if not mr_cols:
                return JSONResponse({"error": "No MR columns selected."}, status_code=400)
            result = tabulate_mr(
                df,
                bundle=mapping_bundle,
                mr_cols=mr_cols,
                col_var=col_var,
                col_var2=None if col_var in {"", "(none)"} or col_var2 == "(none)" else col_var2,
                include_totals=include_totals, 
                include_base=include_base,
                out_counts=include_counts, 
                out_rowpct=include_row_pct, 
                out_colpct=include_col_pct,
                show_row_codes=show_row_codes,
                row_sort_mode="code_asc",
                mr_detection=job.get("mr_detection"),
                custom_groups=_parse_custom_group_definitions(
                    custom_group_definitions,
                    single_label=custom_group_label,
                    single_codes=custom_group_codes,
                ),
                hide_grouped_codes=hide_grouped_codes,
                weight_col=weight_col,
            )
        else:
            if row_var in {"", "(none)"}:
                return JSONResponse({"error": "Select a row variable."}, status_code=400)
            
            if compare_concepts:
                concept_detection = job.get("concept_detection") or {}
                concept_group = (concept_detection.get("member_to_group") or {}).get(row_var)
                if concept_group is None:
                    return JSONResponse({
                        "error": "Concept comparison is only available when the selected row variable belongs to a detected concept family."
                    }, status_code=400)
                result = tabulate_sr_concept_comparison(
                    df,
                    bundle=mapping_bundle,
                    concept_group=concept_group,
                    top_banner_var=None if col_var == "(none)" else col_var,
                    include_totals=include_totals,
                    include_base=include_base,
                    out_counts=include_counts,
                    out_rowpct=include_row_pct,
                    out_colpct=include_col_pct,
                    show_row_codes=show_row_codes,
                    row_sort_mode=row_sort_mode,
                    include_top2_box=include_top2_box,
                    include_bottom2_box=include_bottom2_box,
                    custom_group_label=custom_group_label,
                    custom_group_codes=custom_group_codes,
                    custom_groups=_parse_custom_group_definitions(
                        custom_group_definitions,
                        single_label=custom_group_label,
                        single_codes=custom_group_codes,
                    ),
                    hide_grouped_codes=hide_grouped_codes,
                    weight_col=weight_col,
                )
            else:
                result = tabulate_sr(
                    df,
                    bundle=mapping_bundle,
                    row_var=row_var,
                    col_var=col_var,
                    col_var2=None if col_var in {"", "(none)"} or col_var2 == "(none)" else col_var2,
                    include_totals=include_totals, 
                    include_base=include_base,
                    out_counts=include_counts, 
                    out_rowpct=include_row_pct, 
                    out_colpct=include_col_pct,
                    show_row_codes=show_row_codes,
                    row_sort_mode=row_sort_mode,
                    include_top2_box=include_top2_box,
                    include_bottom2_box=include_bottom2_box,
                    custom_group_label=custom_group_label,
                    custom_group_codes=custom_group_codes,
                    custom_groups=_parse_custom_group_definitions(
                        custom_group_definitions,
                        single_label=custom_group_label,
                        single_codes=custom_group_codes,
                    ),
                    hide_grouped_codes=hide_grouped_codes,
                    weight_col=weight_col,
                )
                result["row_var"] = row_var
        
        result["col_var"] = col_var
        result["col_var2"] = col_var2
        result["mode"] = qtype
        result["weight_col"] = weight_col
        result["show_row_codes"] = show_row_codes
        result["include_top2_box"] = include_top2_box
        result["include_bottom2_box"] = include_bottom2_box
        result["custom_group_definitions"] = custom_group_definitions
        result["custom_group_label"] = custom_group_label
        result["row_sort_mode"] = row_sort_mode
        
        return JSONResponse(result)
        
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)
